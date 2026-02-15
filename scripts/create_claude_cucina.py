"""Crea il file claude_cucina.xlsx con ricette aggiuntive."""
import openpyxl
from openpyxl.styles import Font, Alignment

HEADERS = ["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2", "DIFFICOLTA", "LIBRO_PAG", "PROCEDIMENTO"]

RECIPES = {
    "PRIMI": [
        {
            "RICETTA": "Carbonara classica",
            "INGREDIENTI": "320g di spaghetti, 150g di guanciale, 4 tuorli, 80g di pecorino romano, pepe nero",
            "STAGIONALITA": "All",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare il guanciale a listarelle e rosolare in padella senza olio fino a croccantezza. "
                "2. Sbattere i tuorli con il pecorino grattugiato e abbondante pepe nero. "
                "3. Cuocere gli spaghetti al dente in acqua salata. "
                "4. Scolare la pasta tenendo da parte un po' di acqua di cottura. "
                "5. Versare la pasta nella padella con il guanciale (fuoco spento). "
                "6. Aggiungere la crema di uova e pecorino, mescolare rapidamente aggiungendo acqua di cottura se necessario."
            ),
        },
        {
            "RICETTA": "Cacio e pepe",
            "INGREDIENTI": "320g di tonnarelli, 200g di pecorino romano, pepe nero in grani",
            "STAGIONALITA": "All",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tostare il pepe nero in grani in padella, poi pestare grossolanamente. "
                "2. Grattugiare il pecorino finemente e mescolarlo con un po' di acqua di cottura fino a crema. "
                "3. Cuocere i tonnarelli al dente. "
                "4. Trasferire la pasta in padella con il pepe e un mestolo di acqua di cottura. "
                "5. Mantecare fuori dal fuoco aggiungendo la crema di pecorino a poco a poco. "
                "6. Servire subito con altra grattugiata di pecorino e pepe."
            ),
        },
        {
            "RICETTA": "Pasta e fagioli",
            "INGREDIENTI": "200g di pasta mista corta, 300g di fagioli borlotti, 1 carota, 1 sedano, 1 cipolla, 2 cucchiai di passata di pomodoro, rosmarino, olio, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Soffriggere sedano, carota e cipolla tritati in olio con rosmarino. "
                "2. Aggiungere i fagioli (se secchi, precedentemente ammollati 12 ore) e coprire con acqua. "
                "3. Cuocere per 40 minuti. Frullare metà dei fagioli e rimettere nella pentola. "
                "4. Aggiungere la passata di pomodoro e la pasta. "
                "5. Cuocere fino a che la pasta è pronta, aggiungendo acqua se necessario. "
                "6. Servire con olio a crudo e pepe nero."
            ),
        },
        {
            "RICETTA": "Penne all'arrabbiata",
            "INGREDIENTI": "320g di penne rigate, 400g di pomodori pelati, 2 spicchi d'aglio, peperoncino fresco, prezzemolo, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "glucidica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Soffriggere aglio schiacciato e peperoncino a pezzi in olio abbondante. "
                "2. Aggiungere i pomodori pelati schiacciati con una forchetta. "
                "3. Cuocere il sugo per 15 minuti a fuoco medio, salare. "
                "4. Cuocere le penne al dente e scolarle. "
                "5. Saltare la pasta nel sugo per un minuto. "
                "6. Servire con prezzemolo tritato."
            ),
        },
        {
            "RICETTA": "Minestrone di verdure",
            "INGREDIENTI": "2 patate, 2 zucchine, 2 carote, 1 cipolla, 200g di fagioli cannellini, 100g di pasta corta, pomodoro, basilico, olio, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare tutte le verdure a cubetti piccoli. "
                "2. Soffriggere la cipolla in olio, aggiungere le verdure e i fagioli. "
                "3. Coprire con acqua e cuocere per 30 minuti. "
                "4. Aggiungere la pasta e cuocere fino a cottura. "
                "5. Aggiustare di sale e aggiungere basilico. "
                "6. Servire con un filo d'olio a crudo e parmigiano."
            ),
        },
        {
            "RICETTA": "Risotto ai funghi porcini",
            "INGREDIENTI": "320g di riso carnaroli, 200g di funghi porcini, 1 scalogno, 1 bicchiere di vino bianco, brodo vegetale, burro, parmigiano, prezzemolo, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": None,
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Pulire i funghi porcini e tagliarli a fettine. "
                "2. Soffriggere lo scalogno tritato in burro, aggiungere i funghi e cuocere 5 minuti. "
                "3. Togliere metà dei funghi e tenerli da parte. "
                "4. Aggiungere il riso e tostare per 2 minuti, sfumare con il vino bianco. "
                "5. Aggiungere il brodo caldo un mestolo alla volta, mescolando spesso per 18 minuti. "
                "6. A fine cottura, aggiungere i funghi tenuti da parte, mantecare con burro e parmigiano."
            ),
        },
        {
            "RICETTA": "Pasta al pesto genovese",
            "INGREDIENTI": "320g di trofie, 60g di basilico fresco, 30g di pinoli, 50g di parmigiano, 20g di pecorino, 1 spicchio d'aglio, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "glucidica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Lavare e asciugare il basilico. "
                "2. Nel mortaio (o frullatore), pestare aglio e pinoli con sale grosso. "
                "3. Aggiungere il basilico poco alla volta e pestare fino a crema. "
                "4. Trasferire in una ciotola, aggiungere i formaggi grattugiati e olio a filo. "
                "5. Cuocere le trofie al dente, scolarle tenendo un po' di acqua di cottura. "
                "6. Condire la pasta con il pesto, allungando con acqua di cottura se necessario."
            ),
        },
        {
            "RICETTA": "Zuppa di lenticchie",
            "INGREDIENTI": "300g di lenticchie, 1 carota, 1 sedano, 1 cipolla, 2 pomodori, rosmarino, aglio, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Soffriggere sedano, carota e cipolla tritati con aglio e rosmarino in olio. "
                "2. Aggiungere le lenticchie e i pomodori a pezzi. "
                "3. Coprire con acqua (il doppio del volume delle lenticchie). "
                "4. Cuocere a fuoco basso per 30-40 minuti fino a che le lenticchie sono tenere. "
                "5. Aggiustare di sale e pepe. "
                "6. Servire con crostini di pane e un filo d'olio a crudo."
            ),
        },
    ],
    "SECONDI": [
        {
            "RICETTA": "Scaloppine al marsala",
            "INGREDIENTI": "400g di fettine di vitello, farina, 1 bicchiere di marsala, burro, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Infarinare leggermente le fettine di vitello. "
                "2. Sciogliere il burro in padella e rosolare la carne per 2 minuti per lato. "
                "3. Sfumare con il marsala e lasciar evaporare l'alcol. "
                "4. Cuocere per altri 3-4 minuti fino a che la salsa si addensa. "
                "5. Aggiustare di sale e pepe. "
                "6. Servire con la salsa al marsala e contorno di verdure."
            ),
        },
        {
            "RICETTA": "Branzino al forno",
            "INGREDIENTI": "1 branzino intero (600g), 2 patate, pomodorini, olive, capperi, aglio, rosmarino, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Pulire il branzino, farcire con rosmarino, aglio e limone. "
                "2. Tagliare le patate a fette sottili e disporle in una teglia con olio. "
                "3. Adagiare il branzino sulle patate, aggiungere pomodorini, olive e capperi. "
                "4. Condire con olio, sale e pepe. "
                "5. Cuocere in forno a 200°C per 25-30 minuti. "
                "6. Servire direttamente dalla teglia."
            ),
        },
        {
            "RICETTA": "Involtini di melanzane",
            "INGREDIENTI": "3 melanzane, 250g di ricotta, 100g di prosciutto cotto, 100g di mozzarella, passata di pomodoro, parmigiano, basilico, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare le melanzane a fette lunghe e grigliarle o friggerle. "
                "2. Su ogni fetta adagiare prosciutto, ricotta e mozzarella a dadini. "
                "3. Arrotolare e fissare con uno stecchino. "
                "4. Disporre in una teglia, coprire con passata di pomodoro e parmigiano. "
                "5. Cuocere in forno a 180°C per 20 minuti. "
                "6. Servire caldi con basilico fresco."
            ),
        },
        {
            "RICETTA": "Filetto di salmone in crosta",
            "INGREDIENTI": "4 filetti di salmone, 100g di pangrattato, scorza di limone, aneto, senape, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Mescolare pangrattato con scorza di limone, aneto tritato, olio, sale e pepe. "
                "2. Spennellare i filetti di salmone con senape. "
                "3. Ricoprire con il mix di pangrattato pressando bene. "
                "4. Disporre su una teglia con carta forno. "
                "5. Cuocere in forno a 200°C per 15-18 minuti. "
                "6. Servire con spicchi di limone e insalata."
            ),
        },
        {
            "RICETTA": "Polpettone al forno",
            "INGREDIENTI": "500g di macinato misto, 1 uovo, 50g di pangrattato, 50g di parmigiano, latte, prezzemolo, aglio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Ammollare il pangrattato nel latte. "
                "2. Impastare la carne con uovo, pangrattato strizzato, parmigiano, prezzemolo, aglio tritato, sale e pepe. "
                "3. Formare un filone su un foglio di carta forno. "
                "4. Avvolgere e chiudere le estremità. "
                "5. Cuocere in forno a 180°C per 45-50 minuti, aprendo il cartoccio gli ultimi 10 minuti. "
                "6. Far riposare 10 minuti prima di tagliare a fette."
            ),
        },
        {
            "RICETTA": "Tonno in padella con sesamo",
            "INGREDIENTI": "4 tranci di tonno fresco, 4 cucchiai di semi di sesamo, salsa di soia, zenzero fresco, olio di sesamo, lime",
            "STAGIONALITA": "Estate",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Marinare i tranci di tonno con salsa di soia, zenzero grattugiato e lime per 15 minuti. "
                "2. Ricoprire i tranci con semi di sesamo su entrambi i lati. "
                "3. Scaldare una padella con olio di sesamo a fuoco vivace. "
                "4. Cuocere il tonno per 1-2 minuti per lato (deve restare rosato dentro). "
                "5. Tagliare a fette spesse. "
                "6. Servire con salsa di soia e wasabi."
            ),
        },
        {
            "RICETTA": "Straccetti di manzo con rucola",
            "INGREDIENTI": "400g di fettine di manzo, rucola, pomodorini, scaglie di parmigiano, aceto balsamico, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare le fettine di manzo a strisce. "
                "2. Scaldare una padella con olio a fuoco vivace. "
                "3. Rosolare gli straccetti per 2-3 minuti (devono restare teneri). "
                "4. Disporre su un letto di rucola fresca. "
                "5. Aggiungere pomodorini tagliati a metà e scaglie di parmigiano. "
                "6. Condire con aceto balsamico, olio, sale e pepe."
            ),
        },
    ],
    "PIATTI UNICI": [
        {
            "RICETTA": "Poke bowl con salmone",
            "INGREDIENTI": "200g di salmone fresco, 200g di riso, 1 avocado, edamame, carote, cetriolo, salsa di soia, olio di sesamo, semi di sesamo",
            "STAGIONALITA": "Estate",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Cuocere il riso e farlo raffreddare. "
                "2. Tagliare il salmone a cubetti e marinare con salsa di soia e olio di sesamo. "
                "3. Tagliare avocado, cetriolo e carote a fettine o julienne. "
                "4. Comporre la bowl: riso alla base, poi salmone, verdure e edamame. "
                "5. Condire con salsa di soia e semi di sesamo. "
                "6. Servire fresco."
            ),
        },
        {
            "RICETTA": "Lasagna classica",
            "INGREDIENTI": "250g di lasagne, 300g di macinato di manzo, 500ml di besciamella, 400g di passata, 1 cipolla, 1 carota, 1 sedano, parmigiano, olio, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "difficile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Preparare il ragu: soffriggere sedano, carota e cipolla, aggiungere la carne e rosolare, versare la passata e cuocere 1 ora. "
                "2. Preparare o scaldare la besciamella. "
                "3. In una teglia, alternare strati di lasagne, ragu, besciamella e parmigiano. "
                "4. Terminare con besciamella e parmigiano. "
                "5. Cuocere in forno a 180°C per 30-35 minuti. "
                "6. Far riposare 10 minuti prima di tagliare."
            ),
        },
        {
            "RICETTA": "Buddha bowl con quinoa",
            "INGREDIENTI": "200g di quinoa, 1 patata dolce, 200g di ceci, spinaci, avocado, semi di zucca, tahina, limone, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "fibra",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Cuocere la quinoa in acqua salata per 15 minuti, scolare e raffreddare. "
                "2. Tagliare la patata dolce a cubetti, condire con olio e cuocere in forno a 200°C per 20 minuti. "
                "3. Scaldare i ceci in padella con cumino e paprika. "
                "4. Preparare la salsa: tahina, limone, acqua e sale. "
                "5. Comporre la bowl con quinoa, patata dolce, ceci, spinaci e avocado. "
                "6. Condire con la salsa alla tahina e semi di zucca."
            ),
        },
        {
            "RICETTA": "Frittata di pasta",
            "INGREDIENTI": "200g di spaghetti avanzati, 4 uova, 80g di prosciutto cotto, 80g di mozzarella, parmigiano, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Sbattere le uova con parmigiano, sale e pepe. "
                "2. Tagliare prosciutto e mozzarella a dadini e unire alle uova. "
                "3. Aggiungere la pasta avanzata e mescolare bene. "
                "4. Versare il composto in una padella unta con olio, a fuoco medio-basso. "
                "5. Cuocere con coperchio per 8-10 minuti, poi girare con un piatto. "
                "6. Cuocere l'altro lato per 5 minuti. Servire calda o a temperatura ambiente."
            ),
        },
        {
            "RICETTA": "Wrap di pollo e verdure",
            "INGREDIENTI": "4 tortillas, 300g di petto di pollo, 1 peperone, lattuga, pomodoro, yogurt greco, paprika, cumino, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare il pollo a strisce, condire con paprika, cumino, olio e sale. "
                "2. Cuocere il pollo in padella per 5-6 minuti a fuoco vivace. "
                "3. Tagliare il peperone a striscioline e il pomodoro a fettine. "
                "4. Scaldare le tortillas in padella per 30 secondi per lato. "
                "5. Spalmare yogurt greco sulla tortilla, aggiungere pollo, verdure e lattuga. "
                "6. Arrotolare chiudendo i lati e servire."
            ),
        },
        {
            "RICETTA": "Ribollita toscana",
            "INGREDIENTI": "300g di fagioli cannellini, 1 cavolo nero, 2 fette di pane raffermo, 1 cipolla, 2 carote, 2 pomodori, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "fibra",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Soffriggere cipolla e carote tritate in olio. "
                "2. Aggiungere i fagioli (metà interi, metà frullati) e i pomodori a pezzi. "
                "3. Coprire con acqua e cuocere per 20 minuti. "
                "4. Aggiungere il cavolo nero a striscioline e cuocere altri 15 minuti. "
                "5. Aggiungere il pane raffermo a pezzi, mescolare e cuocere 10 minuti. "
                "6. Servire con olio a crudo e pepe. Si riscalda ancora meglio il giorno dopo."
            ),
        },
        {
            "RICETTA": "Insalata di farro con tonno e pomodorini",
            "INGREDIENTI": "250g di farro, 160g di tonno sott'olio, pomodorini, olive, cipolla rossa, basilico, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Cuocere il farro in acqua salata per 25 minuti, scolare e raffreddare. "
                "2. Sgocciolare il tonno e sminuzzarlo grossolanamente. "
                "3. Tagliare i pomodorini a metà, le olive a rondelle e la cipolla a fettine. "
                "4. Unire tutto in una ciotola capiente. "
                "5. Condire con olio, sale e basilico fresco. "
                "6. Mescolare e servire a temperatura ambiente o freddo."
            ),
        },
    ],
    "CONTORNI": [
        {
            "RICETTA": "Peperonata",
            "INGREDIENTI": "4 peperoni misti, 2 pomodori, 1 cipolla, aglio, basilico, olio, sale, aceto",
            "STAGIONALITA": "Estate",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare i peperoni a listarelle e la cipolla a fettine. "
                "2. Soffriggere la cipolla in olio con aglio. "
                "3. Aggiungere i peperoni e cuocere a fuoco medio per 10 minuti. "
                "4. Aggiungere i pomodori a pezzi, sale e un goccio d'aceto. "
                "5. Cuocere con coperchio per 20 minuti mescolando ogni tanto. "
                "6. Servire con basilico fresco, calda o a temperatura ambiente."
            ),
        },
        {
            "RICETTA": "Patate al rosmarino al forno",
            "INGREDIENTI": "800g di patate, rosmarino, aglio, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "glucidica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Pelare le patate e tagliarle a spicchi regolari. "
                "2. Condire con olio abbondante, rosmarino, aglio schiacciato, sale e pepe. "
                "3. Disporre in una teglia in un singolo strato. "
                "4. Cuocere in forno a 200°C per 35-40 minuti, girando a metà cottura. "
                "5. Devono essere dorate e croccanti fuori, morbide dentro. "
                "6. Servire calde."
            ),
        },
        {
            "RICETTA": "Spinaci saltati con aglio",
            "INGREDIENTI": "500g di spinaci freschi, 2 spicchi d'aglio, peperoncino, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Lavare accuratamente gli spinaci in più acque. "
                "2. Scaldare olio in una padella ampia con aglio a fettine e peperoncino. "
                "3. Aggiungere gli spinaci ancora bagnati e coprire. "
                "4. Cuocere per 3-4 minuti mescolando una volta. "
                "5. Togliere il coperchio, alzare il fuoco e far evaporare l'acqua in eccesso. "
                "6. Salare e servire con un filo d'olio a crudo."
            ),
        },
        {
            "RICETTA": "Caponata siciliana",
            "INGREDIENTI": "3 melanzane, 2 pomodori, 1 cipolla, sedano, olive verdi, capperi, aceto, zucchero, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare le melanzane a cubetti, salare e far scolare 30 minuti. Friggere in olio. "
                "2. Soffriggere cipolla e sedano a pezzi in olio. "
                "3. Aggiungere pomodori a pezzi, olive e capperi. "
                "4. Cuocere 10 minuti, poi aggiungere le melanzane fritte. "
                "5. Versare l'aceto con un cucchiaio di zucchero (agrodolce). "
                "6. Cuocere 5 minuti e servire a temperatura ambiente. Meglio il giorno dopo."
            ),
        },
        {
            "RICETTA": "Fagiolini al pomodoro",
            "INGREDIENTI": "500g di fagiolini, 300g di pomodori pelati, 1 cipolla, aglio, basilico, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Spuntare i fagiolini e lessarli per 5 minuti in acqua salata. Scolare. "
                "2. Soffriggere cipolla e aglio tritati in olio. "
                "3. Aggiungere i pomodori pelati schiacciati e cuocere 10 minuti. "
                "4. Aggiungere i fagiolini e cuocere a fuoco basso per 15 minuti. "
                "5. Aggiustare di sale e aggiungere basilico. "
                "6. Servire caldi o a temperatura ambiente."
            ),
        },
        {
            "RICETTA": "Finocchi gratinati",
            "INGREDIENTI": "4 finocchi, 50g di parmigiano, 30g di pangrattato, burro, noce moscata, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": (
                "1. Tagliare i finocchi a spicchi e lessarli per 10 minuti. Scolare bene. "
                "2. Disporre i finocchi in una teglia imburrata. "
                "3. Mescolare pangrattato con parmigiano, noce moscata, sale e pepe. "
                "4. Cospargere i finocchi con il mix e aggiungere fiocchetti di burro. "
                "5. Cuocere in forno a 200°C per 20 minuti fino a gratinatura dorata. "
                "6. Servire caldi."
            ),
        },
    ],
}


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)

    for sheet_name, recipes in RECIPES.items():
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Write recipes
        for row_idx, recipe in enumerate(recipes, 2):
            for col, header in enumerate(HEADERS, 1):
                ws.cell(row=row_idx, column=col, value=recipe.get(header))

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save("dati/ricette/claude_cucina.xlsx")
    print("File created: dati/ricette/claude_cucina.xlsx")


if __name__ == "__main__":
    main()
