"""Genera ~1000 ricette sane e veloci per il meal planner QuickChef.

Usa generazione combinatoria basata su template con ingredienti curati.
Output: dati/ricette/ricette_1000_healthy.xlsx
"""

import itertools
import random
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Ingredient catalogs
# ---------------------------------------------------------------------------

PASTA = {
    "spaghetti": "320g", "penne": "320g", "fusilli": "320g",
    "rigatoni": "320g", "farfalle": "320g", "orecchiette": "300g",
    "linguine": "320g", "tagliatelle": "300g", "paccheri": "320g",
    "conchiglie": "320g", "casarecce": "320g", "trofie": "300g",
    "bucatini": "320g", "mezze maniche": "320g", "sedanini": "320g",
    "penne integrali": "320g", "fusilli integrali": "320g",
    "spaghetti integrali": "320g",
}

GRAINS = {
    "farro": "250g", "orzo perlato": "250g", "quinoa": "200g",
    "cous-cous": "250g", "bulgur": "200g", "riso integrale": "300g",
    "riso basmati": "300g", "riso carnaroli": "320g", "riso venere": "250g",
    "miglio": "200g",
}

VEGETABLES = {
    "zucchine": "300g", "broccoli": "400g", "melanzane": "400g",
    "peperoni": "300g", "asparagi": "400g", "carciofi": "4",
    "piselli": "200g", "spinaci": "300g", "funghi champignon": "250g",
    "cavolfiore": "400g", "pomodorini": "300g", "fagiolini": "300g",
    "zucca": "400g", "radicchio": "200g", "cicoria": "300g",
    "finocchi": "2", "bietole": "300g", "cime di rapa": "400g",
    "verza": "300g", "cavolo nero": "300g", "sedano rapa": "300g",
    "rape rosse": "300g", "carote": "300g", "porri": "2",
    "cipolla rossa": "1", "topinambur": "300g",
}

FISH = {
    "merluzzo": "400g", "orata": "2 filetti", "branzino": "2 filetti",
    "salmone": "400g", "sgombro": "400g", "pesce spada": "400g",
    "tonno fresco": "400g", "sogliola": "4 filetti", "trota": "2 filetti",
    "sardine": "400g", "gamberi": "400g", "calamari": "400g",
    "cozze": "1kg", "vongole": "1kg", "acciughe": "200g",
    "polpo": "600g",
}

MEAT = {
    "petto di pollo": "400g", "sovracosce di pollo": "600g",
    "fettine di tacchino": "400g", "macinato di tacchino": "400g",
    "fettine di vitello": "400g", "macinato di manzo": "400g",
    "lonza di maiale": "400g", "coniglio": "600g",
}

LEGUMES = {
    "ceci": "300g", "lenticchie": "250g", "fagioli cannellini": "300g",
    "fagioli borlotti": "300g", "fave": "300g", "edamame": "200g",
    "piselli secchi": "200g", "fagioli neri": "300g",
}

HERBS = [
    "basilico", "prezzemolo", "menta", "rosmarino", "timo",
    "origano", "erba cipollina", "salvia", "maggiorana",
]

AROMATICS = [
    "1 spicchio d'aglio", "1/2 cipolla", "1 scalogno",
    "1 cipolla", "1 cipolla rossa", "1 porro",
]

CONDIMENTS = ["olio", "sale", "pepe"]

CHEESES = [
    "parmigiano", "pecorino", "ricotta", "feta", "mozzarella",
    "gorgonzola", "caprino", "burrata", "scamorza",
]

NUTS_SEEDS = [
    "30g di noci", "30g di mandorle", "20g di pinoli",
    "20g di semi di zucca", "20g di semi di sesamo",
    "20g di semi di girasole", "30g di nocciole", "30g di anacardi",
]


# ---------------------------------------------------------------------------
# Template definition
# ---------------------------------------------------------------------------

@dataclass
class RecipeTemplate:
    category: str  # PRIMI, SECONDI, PIATTI UNICI, CONTORNI
    name_fn: object  # callable(*args) -> str
    ingredients_fn: object  # callable(*args) -> str
    slots: dict  # {slot_name: [options]}
    fonte: str
    fonte2: str | None
    stagionalita: str = "All"
    max_generate: int = 50


# ---------------------------------------------------------------------------
# Helper functions for building names & ingredients
# ---------------------------------------------------------------------------

def _join_ingredients(*parts: str) -> str:
    """Join non-empty ingredient parts with commas."""
    return ", ".join(p for p in parts if p)


# ---------------------------------------------------------------------------
# PRIMI templates
# ---------------------------------------------------------------------------

PRIMI_TEMPLATES = [
    # 1. Pasta con verdura singola
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, v: f"{p} con {v}",
        ingredients_fn=lambda p, v: _join_ingredients(
            f"{PASTA[p]} di {p}", f"{VEGETABLES[v]} di {v}",
            random.choice(AROMATICS), "olio, sale, pepe"),
        slots={"p": list(PASTA.keys())[:12], "v": ["zucchine", "broccoli", "asparagi", "carciofi", "funghi champignon", "spinaci"]},
        fonte="glucidica", fonte2="fibra", max_generate=50,
    ),
    # 2. Pasta con verdura e formaggio
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, v, c: f"{p} con {v} e {c}",
        ingredients_fn=lambda p, v, c: _join_ingredients(
            f"{PASTA[p]} di {p}", f"{VEGETABLES[v]} di {v}",
            f"80g di {c}", "olio, sale, pepe"),
        slots={"p": ["penne", "fusilli", "rigatoni", "farfalle", "casarecce", "mezze maniche"],
               "v": ["zucchine", "melanzane", "peperoni", "pomodorini", "radicchio"],
               "c": ["parmigiano", "ricotta", "feta", "pecorino", "caprino"]},
        fonte="glucidica", fonte2="fibra", max_generate=40,
    ),
    # 3. Pasta con pesce
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, f: f"{p} con {f}",
        ingredients_fn=lambda p, f: _join_ingredients(
            f"{PASTA[p]} di {p}", f"{FISH[f]} di {f}",
            "1 spicchio d'aglio", "prezzemolo", "olio, sale, pepe"),
        slots={"p": ["spaghetti", "linguine", "paccheri", "penne", "trofie", "orecchiette"],
               "f": ["gamberi", "vongole", "cozze", "calamari", "tonno fresco", "salmone"]},
        fonte="glucidica", fonte2="proteica", max_generate=30,
    ),
    # 4. Risotto con verdure
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda v: f"Risotto con {v}",
        ingredients_fn=lambda v: _join_ingredients(
            "320g di riso carnaroli", f"{VEGETABLES[v]} di {v}",
            "1 scalogno", "1 bicchiere di vino bianco", "brodo vegetale",
            "burro", "parmigiano", "olio, sale"),
        slots={"v": ["asparagi", "zucchine", "zucca", "radicchio", "carciofi",
                      "funghi champignon", "spinaci", "piselli", "porri",
                      "finocchi", "bietole", "carote"]},
        fonte="glucidica", fonte2="fibra", max_generate=12,
    ),
    # 5. Risotto con pesce
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda f: f"Risotto con {f}",
        ingredients_fn=lambda f: _join_ingredients(
            "320g di riso carnaroli", f"{FISH[f]} di {f}",
            "1 scalogno", "1 bicchiere di vino bianco",
            "prezzemolo", "olio, sale, pepe"),
        slots={"f": ["gamberi", "calamari", "cozze", "vongole", "salmone", "scampi"]},
        fonte="glucidica", fonte2="proteica", max_generate=6,
    ),
    # 6. Insalata di cereali con verdure
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda g, v1, v2: f"Insalata di {g} con {v1} e {v2}",
        ingredients_fn=lambda g, v1, v2: _join_ingredients(
            f"{GRAINS[g]} di {g}", f"{VEGETABLES[v1]} di {v1}",
            f"{VEGETABLES[v2]} di {v2}", "olio, sale, pepe", "basilico"),
        slots={"g": ["farro", "orzo perlato", "quinoa", "cous-cous", "bulgur", "riso venere"],
               "v1": ["pomodorini", "zucchine", "peperoni", "carote", "cetrioli"],
               "v2": ["spinaci", "fagiolini", "radicchio", "finocchi", "rucola"]},
        fonte="glucidica", fonte2="fibra", max_generate=40,
    ),
    # 7. Insalata di cereali con proteina
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda g, pr: f"Insalata di {g} con {pr}",
        ingredients_fn=lambda g, pr: _join_ingredients(
            f"{GRAINS[g]} di {g}", f"200g di {pr}",
            "pomodorini", "cipolla rossa", "olio, sale, pepe"),
        slots={"g": ["farro", "orzo perlato", "quinoa", "cous-cous", "bulgur", "riso integrale"],
               "pr": ["tonno", "pollo", "mozzarella", "feta", "ceci", "fagioli cannellini", "gamberi"]},
        fonte="glucidica", fonte2="proteica", max_generate=35,
    ),
    # 8. Crema/vellutata di verdura
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda v: f"Crema di {v}",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "1 patata", "1/2 cipolla",
            "brodo vegetale", "olio, sale, pepe"),
        slots={"v": ["piselli", "zucchine", "asparagi", "broccoli", "cavolfiore",
                      "carote", "finocchi", "porri", "spinaci", "bietole", "zucca"]},
        fonte="glucidica", fonte2="fibra", max_generate=11,
    ),
    # 9. Pasta al forno
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, v: f"{p} al forno con {v}",
        ingredients_fn=lambda p, v: _join_ingredients(
            f"{PASTA[p]} di {p}", f"{VEGETABLES[v]} di {v}",
            "200g di mozzarella", "passata di pomodoro", "parmigiano",
            "olio, sale"),
        slots={"p": ["penne", "rigatoni", "fusilli", "mezze maniche", "conchiglie", "sedanini"],
               "v": ["melanzane", "zucchine", "peperoni", "broccoli", "spinaci"]},
        fonte="glucidica", fonte2="fibra", max_generate=25,
    ),
    # 10. Pasta con legumi
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, l: f"{p} con {l}",
        ingredients_fn=lambda p, l: _join_ingredients(
            f"{PASTA[p]} di {p}", f"{LEGUMES[l]} di {l}",
            "1 carota", "1 sedano", "1/2 cipolla", "rosmarino", "olio, sale"),
        slots={"p": ["pasta mista corta", "ditalini", "tubetti", "spaghetti spezzati", "orecchiette"],
               "l": ["ceci", "lenticchie", "fagioli cannellini", "fagioli borlotti", "fave"]},
        fonte="glucidica", fonte2="proteica", max_generate=20,
    ),
    # 11. Gnocchi con salsa
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda s: f"Gnocchi {s}",
        ingredients_fn=lambda s: _join_ingredients(
            "500g di gnocchi", f"300g di {s}", "parmigiano", "olio, sale, pepe"),
        slots={"s": ["al pesto", "al pomodoro e basilico", "alla sorrentina",
                      "con gorgonzola e noci", "con zucchine e menta",
                      "con ragù di verdure", "alla crema di zucca",
                      "con broccoli e salsiccia", "con pesto di rucola",
                      "con crema di peperoni", "ai funghi porcini",
                      "alla norma"]},
        fonte="glucidica", fonte2="fibra", max_generate=12,
    ),
    # 12. Pasta con due verdure
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, v1, v2: f"{p} con {v1} e {v2}",
        ingredients_fn=lambda p, v1, v2: _join_ingredients(
            f"{PASTA[p]} di {p}", f"200g di {v1}", f"200g di {v2}",
            "1 spicchio d'aglio", "olio, sale, pepe"),
        slots={"p": ["penne", "fusilli", "orecchiette", "casarecce", "farfalle", "rigatoni",
                      "spaghetti", "linguine", "trofie"],
               "v1": ["zucchine", "melanzane", "peperoni", "pomodorini", "broccoli"],
               "v2": ["olive", "capperi", "pinoli", "mandorle", "noci"]},
        fonte="glucidica", fonte2="fibra", max_generate=40,
    ),
    # 13. Zuppa calda
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda v1, v2: f"Zuppa di {v1} e {v2}",
        ingredients_fn=lambda v1, v2: _join_ingredients(
            f"{VEGETABLES[v1]} di {v1}", f"{VEGETABLES[v2]} di {v2}",
            "1 cipolla", "1 carota", "brodo vegetale", "olio, sale, pepe"),
        slots={"v1": ["zucca", "carote", "patate", "cavolo nero", "verza", "porri"],
               "v2": ["patate", "piselli", "lenticchie", "fagioli", "ceci", "farro"]},
        fonte="glucidica", fonte2="fibra", stagionalita="Inverno", max_generate=25,
    ),
    # 14. Pasta fredda estiva
    RecipeTemplate(
        category="PRIMI",
        name_fn=lambda p, v: f"{p} freddi con {v}",
        ingredients_fn=lambda p, v: _join_ingredients(
            f"{PASTA[p]} di {p}", f"200g di {v}",
            "pomodorini", "basilico", "olio, sale"),
        slots={"p": ["fusilli", "penne", "farfalle", "conchiglie", "sedanini",
                      "fusilli integrali", "penne integrali"],
               "v": ["mozzarella", "tonno", "olive e capperi",
                      "rucola e parmigiano", "verdure grigliate",
                      "pesto e fagiolini"]},
        fonte="glucidica", fonte2="fibra", stagionalita="Estate", max_generate=30,
    ),
]

# ---------------------------------------------------------------------------
# SECONDI templates
# ---------------------------------------------------------------------------

SECONDI_TEMPLATES = [
    # 1. Pesce al forno/padella
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda f, m: f"{f} {m}",
        ingredients_fn=lambda f, m: _join_ingredients(
            f"{FISH[f]} di {f}", "1 limone", random.choice(["prezzemolo", "timo", "rosmarino"]),
            "olio, sale, pepe"),
        slots={"f": ["merluzzo", "orata", "branzino", "salmone", "sgombro",
                      "pesce spada", "sogliola", "trota", "sardine"],
               "m": ["al forno", "in padella", "alla griglia", "al cartoccio",
                      "al vapore", "gratinato"]},
        fonte="proteica", fonte2=None, max_generate=45,
    ),
    # 2. Pesce con verdure
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda f, v: f"{f} con {v}",
        ingredients_fn=lambda f, v: _join_ingredients(
            f"{FISH[f]} di {f}", f"{VEGETABLES[v]} di {v}",
            "1 limone", "olio, sale, pepe"),
        slots={"f": ["salmone", "merluzzo", "orata", "branzino", "pesce spada",
                      "tonno fresco", "sgombro", "trota"],
               "v": ["pomodorini", "spinaci", "asparagi", "fagiolini",
                      "zucchine", "peperoni", "carciofi", "finocchi"]},
        fonte="proteica", fonte2="fibra", max_generate=45,
    ),
    # 3. Pollo variazioni
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda s: f"Pollo {s}",
        ingredients_fn=lambda s: _join_ingredients(
            "400g di petto di pollo", f"{s}", "olio, sale, pepe"),
        slots={"s": [
            "al limone e rosmarino", "alla paprika", "al curry e latte di cocco",
            "alla cacciatora", "con peperoni", "alle erbe aromatiche",
            "alla senape e miele", "al sesamo", "alla curcuma",
            "marinato allo yogurt", "con olive e capperi", "alla salvia",
            "con funghi trifolati", "alla pizzaiola", "al marsala",
            "con verdure croccanti", "alla griglia con limone",
            "allo zenzero e soia", "con spinaci e pinoli",
            "al forno con patate", "con pomodorini e basilico",
        ]},
        fonte="proteica", fonte2=None, max_generate=21,
    ),
    # 4. Tacchino variazioni
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda s: f"Tacchino {s}",
        ingredients_fn=lambda s: _join_ingredients(
            "400g di fettine di tacchino", f"{s}", "olio, sale, pepe"),
        slots={"s": [
            "al limone", "alla senape", "con verdure miste", "alla griglia",
            "al forno con rosmarino", "con funghi", "alla salvia e burro",
            "alla pizzaiola", "con peperoni e olive", "con zucchine",
            "al curry", "marinato alle erbe", "con spinaci",
            "alla paprika affumicata", "con carote glassate",
        ]},
        fonte="proteica", fonte2=None, max_generate=15,
    ),
    # 5. Frittate
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda v: f"Frittata di {v}",
        ingredients_fn=lambda v: _join_ingredients(
            "4 uova", f"{VEGETABLES[v]} di {v}",
            "parmigiano", "olio, sale, pepe"),
        slots={"v": ["zucchine", "spinaci", "asparagi", "peperoni", "porri",
                      "carciofi", "cipolle", "bietole", "funghi champignon",
                      "finocchi", "fagiolini", "cime di rapa", "piselli",
                      "radicchio", "zucca"]},
        fonte="proteica", fonte2="fibra", max_generate=15,
    ),
    # 6. Polpette varie
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda p: f"Polpette di {p}",
        ingredients_fn=lambda p: _join_ingredients(
            f"400g di {p}", "1 uovo", "30g di pangrattato",
            "parmigiano", "prezzemolo", "olio, sale, pepe"),
        slots={"p": ["pollo", "tacchino", "merluzzo", "tonno", "salmone",
                      "manzo", "vitello", "ceci", "lenticchie",
                      "fagioli cannellini", "fagioli neri", "melanzane",
                      "zucchine e ricotta", "patate e formaggio"]},
        fonte="proteica", fonte2=None, max_generate=14,
    ),
    # 7. Carne rossa
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda t, s: f"{t} {s}",
        ingredients_fn=lambda t, s: _join_ingredients(
            f"400g di {t}", f"{s}", "olio, sale, pepe"),
        slots={"t": ["tagliata di manzo", "fettine di vitello", "scaloppine",
                      "straccetti di manzo", "hamburger di manzo", "macinato di manzo"],
               "s": ["con rucola e parmigiano", "al limone", "alla griglia",
                      "con funghi porcini", "al rosmarino", "alla senape",
                      "con pomodorini", "all'aceto balsamico"]},
        fonte="proteica", fonte2=None, max_generate=35,
    ),
    # 8. Burger di legumi
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda l: f"Burger di {l}",
        ingredients_fn=lambda l: _join_ingredients(
            f"{LEGUMES[l]} di {l}", "50g di pangrattato",
            "1/2 cipolla", "1 cucchiaio di salsa di soia",
            "cumino", "olio, sale, pepe"),
        slots={"l": ["ceci", "lenticchie", "fagioli cannellini", "fagioli borlotti",
                      "fagioli neri", "fave", "edamame"]},
        fonte="proteica", fonte2="glucidica", max_generate=7,
    ),
    # 9. Crostacei e molluschi
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda f, s: f"{f} {s}",
        ingredients_fn=lambda f, s: _join_ingredients(
            f"{FISH[f]} di {f}", f"{s}",
            "prezzemolo", "olio, sale, pepe"),
        slots={"f": ["gamberi", "calamari", "polpo", "cozze"],
               "s": ["alla griglia", "in padella con aglio",
                      "al forno", "alla marinara", "con pomodorini",
                      "alla piastra con limone"]},
        fonte="proteica", fonte2=None, max_generate=20,
    ),
    # 10. Uova preparate
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda s: f"Uova {s}",
        ingredients_fn=lambda s: _join_ingredients(
            "4 uova", f"{s}", "olio, sale, pepe"),
        slots={"s": [
            "in camicia su spinaci", "strapazzate con pomodorini",
            "al forno con peperoni", "alla fiorentina",
            "con asparagi e parmigiano", "in purgatorio",
            "strapazzate con funghi", "al tegamino con tartufo",
            "sode con salsa tonnata", "con zucchine trifolate",
        ]},
        fonte="proteica", fonte2="fibra", max_generate=10,
    ),
    # 11. Pesce in crosta/impanato
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda f: f"{f} in crosta di erbe",
        ingredients_fn=lambda f: _join_ingredients(
            f"{FISH[f]} di {f}", "50g di pangrattato",
            "prezzemolo", "1 limone", "olio, sale, pepe"),
        slots={"f": ["salmone", "merluzzo", "orata", "branzino",
                      "pesce spada", "sgombro", "trota", "sogliola"]},
        fonte="proteica", fonte2=None, max_generate=8,
    ),
    # 12. Maiale leggero
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda s: f"Lonza di maiale {s}",
        ingredients_fn=lambda s: _join_ingredients(
            "400g di lonza di maiale", f"{s}", "olio, sale, pepe"),
        slots={"s": [
            "al forno con mele", "alla senape", "con funghi",
            "con peperoni", "marinata alle erbe", "al rosmarino",
            "con salsa di soia e zenzero", "alla griglia",
            "con prugne e timo", "con patate al forno",
        ]},
        fonte="proteica", fonte2=None, max_generate=10,
    ),
    # 13. Tofu/tempeh
    RecipeTemplate(
        category="SECONDI",
        name_fn=lambda p, s: f"{p} {s}",
        ingredients_fn=lambda p, s: _join_ingredients(
            f"300g di {p}", f"{s}", "salsa di soia", "olio, sale, pepe"),
        slots={"p": ["tofu", "tempeh"],
               "s": ["alla piastra con verdure", "marinato alla soia",
                      "con funghi e sesamo", "saltato con broccoli",
                      "al curry", "con peperoni e zenzero",
                      "croccante con salsa agrodolce"]},
        fonte="proteica", fonte2="fibra", max_generate=14,
    ),
]

# ---------------------------------------------------------------------------
# PIATTI UNICI templates
# ---------------------------------------------------------------------------

PIATTI_UNICI_TEMPLATES = [
    # 1. Bowl con proteina e cereale
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda g, pr: f"Bowl di {g} con {pr}",
        ingredients_fn=lambda g, pr: _join_ingredients(
            f"{GRAINS[g]} di {g}", f"200g di {pr}",
            "1 avocado", "carote", "cetriolo", "olio, sale"),
        slots={"g": ["riso integrale", "quinoa", "riso venere", "farro", "bulgur", "cous-cous"],
               "pr": ["salmone", "tonno", "pollo", "tofu", "gamberi", "ceci",
                       "edamame", "uova sode", "feta"]},
        fonte="proteica", fonte2="glucidica", max_generate=45,
    ),
    # 2. Insalatona completa
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda pr, v: f"Insalatona con {pr} e {v}",
        ingredients_fn=lambda pr, v: _join_ingredients(
            "lattuga", f"200g di {pr}", f"200g di {v}",
            "pomodorini", "olio, sale, pepe"),
        slots={"pr": ["pollo grigliato", "tonno", "uova sode", "salmone affumicato",
                       "mozzarella", "feta", "gamberi", "ceci", "fagioli cannellini"],
               "v": ["avocado", "carote", "mais", "cetrioli", "peperoni",
                      "olive", "noci", "semi di zucca"]},
        fonte="proteica", fonte2="fibra", max_generate=45,
    ),
    # 3. Proteina con cereale e verdura
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda pr, g, v: f"{pr} con {g} e {v}",
        ingredients_fn=lambda pr, g, v: _join_ingredients(
            f"300g di {pr}", f"{GRAINS[g]} di {g}",
            f"200g di {v}", "olio, sale, pepe"),
        slots={"pr": ["pollo", "salmone", "tacchino", "merluzzo", "tonno"],
               "g": ["riso integrale", "quinoa", "farro", "cous-cous", "bulgur"],
               "v": ["broccoli", "spinaci", "zucchine", "fagiolini", "peperoni"]},
        fonte="proteica", fonte2="glucidica", max_generate=45,
    ),
    # 4. Zuppe complete con legumi
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda l, v: f"Zuppa di {l} e {v}",
        ingredients_fn=lambda l, v: _join_ingredients(
            f"{LEGUMES[l]} di {l}", f"{VEGETABLES[v]} di {v}",
            "1 cipolla", "1 carota", "brodo vegetale", "olio, sale, pepe"),
        slots={"l": ["ceci", "lenticchie", "fagioli cannellini", "fagioli borlotti",
                      "fave", "fagioli neri"],
               "v": ["cavolo nero", "spinaci", "zucca", "verza",
                      "bietole", "carote", "porri", "sedano rapa"]},
        fonte="proteica", fonte2="fibra", stagionalita="Inverno", max_generate=35,
    ),
    # 5. Wrap/Piadine
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda pr, v: f"Wrap con {pr} e {v}",
        ingredients_fn=lambda pr, v: _join_ingredients(
            "2 tortillas integrali", f"200g di {pr}", f"150g di {v}",
            "yogurt greco", "olio, sale"),
        slots={"pr": ["pollo grigliato", "tacchino", "tonno", "feta",
                       "hummus", "salmone affumicato", "uova strapazzate"],
               "v": ["lattuga e pomodorini", "verdure grigliate", "avocado e carote",
                      "spinaci e pomodori secchi", "rucola e parmigiano"]},
        fonte="proteica", fonte2="glucidica", max_generate=25,
    ),
    # 6. Piatti one-pot con cereale
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda g, pr: f"{g} con {pr}",
        ingredients_fn=lambda g, pr: _join_ingredients(
            f"{GRAINS[g]} di {g}", f"300g di {pr}",
            "1 cipolla", "1 carota", "olio, sale, pepe"),
        slots={"g": ["riso integrale", "farro", "orzo perlato", "quinoa", "cous-cous", "bulgur"],
               "pr": ["pollo e verdure miste", "salsiccia e broccoli",
                       "ceci e spinaci", "lenticchie e pomodoro",
                       "gamberi e zucchine", "tacchino e peperoni",
                       "tofu e funghi"]},
        fonte="glucidica", fonte2="proteica", max_generate=35,
    ),
    # 7. Vellutate proteiche
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda l: f"Vellutata di {l}",
        ingredients_fn=lambda l: _join_ingredients(
            f"{LEGUMES[l]} di {l}", "1 patata", "1 cipolla",
            "1 carota", "brodo vegetale", "olio, sale, pepe"),
        slots={"l": ["lenticchie rosse", "ceci", "piselli", "fave",
                      "fagioli cannellini", "lenticchie e zucca",
                      "ceci e carote", "fagioli e cavolo nero"]},
        fonte="proteica", fonte2="fibra", stagionalita="Inverno", max_generate=8,
    ),
    # 8. Piatti unici mediterranei
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda n: n,
        ingredients_fn=lambda n: {
            "Shakshuka con uova e peperoni": "4 uova, 400g di pomodori pelati, 2 peperoni, 1 cipolla, cumino, paprika, olio, sale",
            "Moussaka leggera": "2 melanzane, 300g di macinato di manzo, 200ml di besciamella, pomodoro, parmigiano, olio, sale",
            "Tortilla spagnola di patate": "4 uova, 3 patate, 1 cipolla, olio, sale, pepe",
            "Tabbouleh con ceci": "200g di bulgur, 200g di ceci, pomodorini, cetriolo, prezzemolo, menta, limone, olio, sale",
            "Falafel con insalata e hummus": "300g di ceci, 1 cipolla, prezzemolo, cumino, insalata mista, tahina, olio, sale",
            "Ratatouille con uova al forno": "1 zucchina, 1 melanzana, 1 peperone, pomodorini, 4 uova, timo, olio, sale",
            "Tajine di pollo e verdure": "400g di pollo, 2 carote, 1 zucchina, olive, limone confit, cumino, olio, sale",
            "Pita con falafel e verdure": "300g di ceci, 4 pita integrali, lattuga, pomodoro, cipolla, tahina, olio, sale",
            "Insalata greca con farro": "250g di farro, 200g di feta, cetriolo, pomodorini, olive, cipolla rossa, origano, olio",
            "Paella vegetariana": "300g di riso, 200g di fagiolini, carciofi, peperoni, piselli, zafferano, olio, sale",
            "Couscous con agnello e verdure": "250g di cous-cous, 300g di agnello, 2 carote, 1 zucchina, ceci, cumino, olio, sale",
            "Bibimbap vegetariano": "300g di riso, 200g di spinaci, carote, zucchine, funghi, uovo, salsa di soia, olio di sesamo",
            "Frittata di quinoa e verdure": "200g di quinoa, 4 uova, spinaci, pomodorini, parmigiano, olio, sale",
            "Polenta con ragù di funghi": "300g di polenta, 400g di funghi misti, 1 cipolla, timo, parmigiano, olio, sale",
            "Pasta e ceci": "200g di pasta corta, 300g di ceci, rosmarino, 1 spicchio d'aglio, pomodoro, olio, sale",
            "Minestra di farro e fagioli": "200g di farro, 300g di fagioli borlotti, 1 carota, 1 sedano, 1 cipolla, rosmarino, olio, sale",
        }.get(n, ""),
        slots={"n": [
            "Shakshuka con uova e peperoni", "Moussaka leggera",
            "Tortilla spagnola di patate", "Tabbouleh con ceci",
            "Falafel con insalata e hummus", "Ratatouille con uova al forno",
            "Tajine di pollo e verdure", "Pita con falafel e verdure",
            "Insalata greca con farro", "Paella vegetariana",
            "Couscous con agnello e verdure", "Bibimbap vegetariano",
            "Frittata di quinoa e verdure", "Polenta con ragù di funghi",
            "Pasta e ceci", "Minestra di farro e fagioli",
        ]},
        fonte="proteica", fonte2="glucidica", max_generate=16,
    ),
    # 9. Pesce con patate
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda f: f"{f} con patate al forno",
        ingredients_fn=lambda f: _join_ingredients(
            f"{FISH[f]} di {f}", "400g di patate", "pomodorini",
            "rosmarino", "olio, sale, pepe"),
        slots={"f": ["salmone", "merluzzo", "orata", "branzino",
                      "pesce spada", "sgombro", "trota", "sogliola"]},
        fonte="proteica", fonte2="glucidica", max_generate=8,
    ),
    # 10. Pollo/tacchino con cereale e verdura
    RecipeTemplate(
        category="PIATTI UNICI",
        name_fn=lambda m, v: f"Pollo {m} con {v}",
        ingredients_fn=lambda m, v: _join_ingredients(
            "400g di petto di pollo", f"{m}",
            f"200g di {v}", "olio, sale, pepe"),
        slots={"m": ["alla griglia", "al forno", "in padella", "marinato"],
               "v": ["insalata di quinoa", "riso basmati e spinaci",
                      "cous-cous e verdure", "farro e pomodorini",
                      "patate dolci al forno", "bulgur e zucchine"]},
        fonte="proteica", fonte2="glucidica", max_generate=20,
    ),
]

# ---------------------------------------------------------------------------
# CONTORNI templates
# ---------------------------------------------------------------------------

CONTORNI_TEMPLATES = [
    # 1. Verdure al forno
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} al forno",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "olio, sale, pepe"),
        slots={"v": ["zucchine", "melanzane", "peperoni", "cavolfiore",
                      "finocchi", "carciofi", "zucca", "carote",
                      "fagiolini", "pomodorini", "broccoli", "porri",
                      "bietole", "rape rosse", "topinambur"]},
        fonte="fibra", fonte2=None, max_generate=15,
    ),
    # 2. Verdure in padella
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} in padella",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "1 spicchio d'aglio",
            "olio, sale, pepe"),
        slots={"v": ["spinaci", "bietole", "cicoria", "cime di rapa",
                      "broccoli", "fagiolini", "zucchine", "peperoni",
                      "funghi champignon", "cavolo nero", "verza",
                      "asparagi", "piselli", "porri"]},
        fonte="fibra", fonte2=None, max_generate=14,
    ),
    # 3. Verdure grigliate
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} alla griglia",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "olio, sale, pepe", "aceto balsamico"),
        slots={"v": ["zucchine", "melanzane", "peperoni", "radicchio",
                      "finocchi", "asparagi", "porri", "cipolla rossa"]},
        fonte="fibra", fonte2=None, max_generate=8,
    ),
    # 4. Verdure al vapore
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} al vapore",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "1 limone", "olio, sale"),
        slots={"v": ["broccoli", "cavolfiore", "fagiolini", "carote",
                      "zucchine", "asparagi", "spinaci", "bietole",
                      "finocchi", "piselli"]},
        fonte="fibra", fonte2=None, max_generate=10,
    ),
    # 5. Verdure gratinate
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} gratinati",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "30g di pangrattato",
            "parmigiano", "olio, sale"),
        slots={"v": ["finocchi", "zucchine", "cavolfiore", "porri",
                      "carciofi", "asparagi", "melanzane", "pomodorini",
                      "sedano rapa", "carote"]},
        fonte="fibra", fonte2=None, max_generate=10,
    ),
    # 6. Insalate crude
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v1, v2: f"Insalata di {v1} e {v2}",
        ingredients_fn=lambda v1, v2: _join_ingredients(
            f"200g di {v1}", f"200g di {v2}", "olio, sale, pepe"),
        slots={"v1": ["finocchi", "carote", "radicchio", "sedano", "cetrioli",
                       "cavolo cappuccio", "barbabietola", "rucola"],
               "v2": ["arance", "mele", "noci", "parmigiano", "olive",
                       "melograno", "avocado", "semi di girasole"]},
        fonte="fibra", fonte2=None, max_generate=35,
    ),
    # 7. Patate variazioni
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda s: f"Patate {s}",
        ingredients_fn=lambda s: _join_ingredients(
            "500g di patate", f"{s}", "olio, sale, pepe"),
        slots={"s": [
            "al forno con paprika", "al rosmarino", "novelle al timo",
            "schiacciate con aglio", "dolci al forno", "lesse con prezzemolo",
            "alla curcuma", "con olive e capperi", "al forno con senape",
            "novelle al forno con erbe", "alla provenzale",
            "con cipolla e origano", "alla diavola",
        ]},
        fonte="glucidica", fonte2=None, max_generate=13,
    ),
    # 8. Legumi come contorno
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda l, s: f"{l} {s}",
        ingredients_fn=lambda l, s: _join_ingredients(
            f"{LEGUMES[l]} di {l}", f"{s}", "olio, sale, pepe"),
        slots={"l": ["fagioli cannellini", "ceci", "lenticchie",
                      "fagioli borlotti", "fave", "edamame"],
               "s": ["all'uccelletto", "in insalata", "con cipolla e prezzemolo",
                      "al pomodoro", "al rosmarino", "con erbe aromatiche"]},
        fonte="fibra", fonte2="proteica", max_generate=25,
    ),
    # 9. Verdure ripiene
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} ripieni",
        ingredients_fn=lambda v: _join_ingredients(
            f"4 {v}", "100g di pangrattato", "50g di parmigiano",
            "prezzemolo", "1 spicchio d'aglio", "olio, sale, pepe"),
        slots={"v": ["peperoni", "zucchine", "pomodori", "melanzane",
                      "cipolle", "funghi champignon"]},
        fonte="fibra", fonte2=None, max_generate=6,
    ),
    # 10. Verdure in umido
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"{v} in umido",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "200g di pomodori pelati",
            "1/2 cipolla", "olio, sale, pepe"),
        slots={"v": ["fagiolini", "carciofi", "zucchine", "peperoni",
                      "melanzane", "piselli", "bietole", "verza",
                      "cavolo nero", "finocchi"]},
        fonte="fibra", fonte2=None, max_generate=10,
    ),
    # 11. Verdure al forno con condimento speciale
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v, c: f"{v} al forno con {c}",
        ingredients_fn=lambda v, c: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", f"{c}", "olio, sale, pepe"),
        slots={"v": ["cavolfiore", "broccoli", "zucca", "carote",
                      "finocchi", "patate dolci", "rape rosse"],
               "c": ["tahina e limone", "miele e timo", "paprika affumicata",
                      "semi di sesamo", "parmigiano e pangrattato",
                      "curry e curcuma", "rosmarino e aglio"]},
        fonte="fibra", fonte2=None, max_generate=30,
    ),
    # 12. Purè/crema come contorno
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"Purè di {v}",
        ingredients_fn=lambda v: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", "1 patata",
            "burro", "noce moscata", "sale"),
        slots={"v": ["piselli", "carote", "zucca", "cavolfiore",
                      "sedano rapa", "finocchi", "broccoli", "spinaci"]},
        fonte="fibra", fonte2=None, max_generate=8,
    ),
    # 13. Chips di verdure
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v: f"Chips di {v} al forno",
        ingredients_fn=lambda v: _join_ingredients(
            f"400g di {v}", "olio, sale, paprika"),
        slots={"v": ["zucchine", "barbabietola", "patate dolci",
                      "cavolo nero", "carote", "sedano rapa"]},
        fonte="fibra", fonte2=None, max_generate=6,
    ),
    # 14. Verdure stufate con erbe
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda v, h: f"{v} con {h}",
        ingredients_fn=lambda v, h: _join_ingredients(
            f"{VEGETABLES[v]} di {v}", f"{h}",
            "1 spicchio d'aglio", "olio, sale, pepe"),
        slots={"v": ["zucchine", "carote", "finocchi", "porri",
                      "rape rosse", "cavolo nero", "sedano rapa"],
               "h": ["menta fresca", "prezzemolo e limone",
                      "timo e aglio", "rosmarino", "salvia e burro"]},
        fonte="fibra", fonte2=None, max_generate=25,
    ),
    # 15. Contorni specifici italiani
    RecipeTemplate(
        category="CONTORNI",
        name_fn=lambda n: n,
        ingredients_fn=lambda n: {
            "Panzanella toscana": "200g di pane raffermo, pomodorini, cetriolo, cipolla rossa, basilico, olio, aceto, sale",
            "Cipolline in agrodolce": "400g di cipolline borettane, aceto balsamico, 1 cucchiaio di zucchero, olio, sale",
            "Carciofi alla romana": "4 carciofi, prezzemolo, mentuccia, 1 spicchio d'aglio, olio, sale, pepe",
            "Parmigiana di melanzane": "3 melanzane, passata di pomodoro, mozzarella, parmigiano, basilico, olio, sale",
            "Asparagi alla milanese": "500g di asparagi, 2 uova, parmigiano, burro, sale, pepe",
            "Catalogna ripassata": "500g di catalogna, 1 spicchio d'aglio, peperoncino, olio, sale",
            "Friarielli napoletani": "500g di friarielli, 2 spicchi d'aglio, peperoncino, olio, sale",
            "Carciofi trifolati": "4 carciofi, prezzemolo, 1 spicchio d'aglio, olio, sale, pepe",
            "Verza stufata con mele": "300g di verza, 1 mela, 1/2 cipolla, aceto di mele, olio, sale",
            "Scarola in padella": "1 scarola, olive, capperi, 1 spicchio d'aglio, olio, sale",
            "Broccoletti ripassati": "400g di broccoletti, 1 spicchio d'aglio, peperoncino, olio, sale",
            "Cavolini di Bruxelles al forno": "400g di cavolini di Bruxelles, 30g di pancetta, olio, sale, pepe",
        }.get(n, ""),
        slots={"n": [
            "Panzanella toscana", "Cipolline in agrodolce",
            "Carciofi alla romana", "Parmigiana di melanzane",
            "Asparagi alla milanese", "Catalogna ripassata",
            "Friarielli napoletani", "Carciofi trifolati",
            "Verza stufata con mele", "Scarola in padella",
            "Broccoletti ripassati", "Cavolini di Bruxelles al forno",
        ]},
        fonte="fibra", fonte2=None, max_generate=12,
    ),
]


# ---------------------------------------------------------------------------
# Generator
# ---------------------------------------------------------------------------

class RecipeGenerator:
    def __init__(self):
        self._used_names: set[str] = set()
        self._existing_names: set[str] = set()

    def set_existing_names(self, names: set[str]):
        """Exclude names from existing recipe files."""
        self._existing_names = {n.strip().lower() for n in names}

    def generate_all(
        self,
        target_per_category: dict[str, int] | None = None,
    ) -> dict[str, list[dict]]:
        """Generate recipes from all templates."""
        if target_per_category is None:
            target_per_category = {
                "PRIMI": 300, "SECONDI": 250,
                "PIATTI UNICI": 250, "CONTORNI": 200,
            }

        all_templates = {
            "PRIMI": PRIMI_TEMPLATES,
            "SECONDI": SECONDI_TEMPLATES,
            "PIATTI UNICI": PIATTI_UNICI_TEMPLATES,
            "CONTORNI": CONTORNI_TEMPLATES,
        }

        result: dict[str, list[dict]] = {cat: [] for cat in target_per_category}

        for category, templates in all_templates.items():
            target = target_per_category[category]
            for tmpl in templates:
                if len(result[category]) >= target:
                    break
                recipes = self._expand_template(tmpl)
                remaining = target - len(result[category])
                result[category].extend(recipes[:remaining])

        return result

    def _expand_template(self, tmpl: RecipeTemplate) -> list[dict]:
        """Generate recipes from a single template."""
        slot_names = list(tmpl.slots.keys())
        slot_values = [tmpl.slots[k] for k in slot_names]

        combinations = list(itertools.product(*slot_values))
        random.shuffle(combinations)

        recipes = []
        for combo in combinations:
            if len(recipes) >= tmpl.max_generate:
                break

            args = dict(zip(slot_names, combo))
            try:
                name = tmpl.name_fn(*combo)
                ingredients = tmpl.ingredients_fn(*combo)
            except Exception:
                continue

            # Skip if no ingredients generated
            if not ingredients:
                continue

            # Capitalize first letter
            name = name[0].upper() + name[1:] if name else name

            # Check uniqueness
            name_lower = name.strip().lower()
            if name_lower in self._used_names or name_lower in self._existing_names:
                continue

            self._used_names.add(name_lower)
            recipes.append({
                "RICETTA": name,
                "INGREDIENTI": ingredients,
                "STAGIONALITA": tmpl.stagionalita,
                "FONTE": tmpl.fonte,
                "FONTE 2": tmpl.fonte2,
            })

        return recipes


# ---------------------------------------------------------------------------
# Excel Writer
# ---------------------------------------------------------------------------

HEADERS = ["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"]


def write_excel(recipes: dict[str, list[dict]], filepath: str):
    """Write recipes to an Excel file with 4 sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)

    for sheet_name in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Write recipes
        for row_idx, recipe in enumerate(recipes.get(sheet_name, []), 2):
            for col, header in enumerate(HEADERS, 1):
                ws.cell(row=row_idx, column=col, value=recipe.get(header))

        # Auto-width columns
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 80)

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import sys
    sys.path.insert(0, ".")

    output_path = "dati/ricette/ricette_1000_healthy.xlsx"

    # Remove old generated file before loading existing names
    import os
    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"Removed old {output_path}")

    # Load existing recipe names to avoid duplicates
    existing_names: set[str] = set()
    try:
        from src.excel import RecipeRepository
        repo = RecipeRepository("dati/ricette")
        by_cat, _ = repo.load_recipes()
        for cat_recipes in by_cat.values():
            for r in cat_recipes:
                existing_names.add(r.name.strip().lower())
        print(f"Loaded {len(existing_names)} existing recipe names to avoid duplicates")
    except Exception as e:
        print(f"Warning: could not load existing recipes: {e}")

    # Generate
    gen = RecipeGenerator()
    gen.set_existing_names(existing_names)
    recipes = gen.generate_all()

    # Stats
    total = 0
    for cat, recs in recipes.items():
        print(f"  {cat}: {len(recs)} ricette")
        total += len(recs)
    print(f"  TOTALE: {total} ricette")

    # Write
    write_excel(recipes, output_path)
    print(f"\nFile scritto: {output_path}")

    # Validate by re-loading
    try:
        from src.excel import RecipeRepository
        repo2 = RecipeRepository("dati/ricette")
        by_cat2, _ = repo2.load_recipes()
        total_loaded = sum(len(v) for v in by_cat2.values())
        print(f"\nValidazione: {total_loaded} ricette totali caricate (incluse esistenti)")
        for cat, recs in by_cat2.items():
            print(f"  {cat}: {len(recs)}")
    except Exception as e:
        print(f"Errore di validazione: {e}")


if __name__ == "__main__":
    main()
