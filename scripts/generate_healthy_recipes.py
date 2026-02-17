"""Genera un file Excel di ricette healthy e sazianti."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

OUTPUT = Path(__file__).resolve().parent.parent / "dati" / "ricette" / "healthy_sazianti.xlsx"

COLUMNS = ["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2",
           "DIFFICOLTA", "LIBRO_PAG", "PROCEDIMENTO"]

RECIPES = {
    "PRIMI": [
        {
            "RICETTA": "Zuppa di lenticchie rosse e curcuma",
            "INGREDIENTI": "250g di lenticchie rosse, 1 cipolla, 2 carote, 1 patata, 1 cucchiaino di curcuma, 1 cucchiaino di cumino, zenzero fresco, 1 litro di brodo vegetale, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "proteica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Soffriggere cipolla tritata in poco olio. 2. Aggiungere carote e patata a dadini, curcuma, cumino e zenzero grattugiato. 3. Unire le lenticchie rosse e il brodo. 4. Cuocere per 20-25 minuti fino a che le lenticchie si disfano. 5. Frullare parzialmente per ottenere una consistenza cremosa. 6. Servire con un filo d'olio a crudo e pepe nero.",
        },
        {
            "RICETTA": "Pasta integrale con broccoli e acciughe",
            "INGREDIENTI": "320g di penne integrali, 400g di broccoli, 4 filetti di acciuga, 1 spicchio d'aglio, peperoncino, olio, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Lessare i broccoli divisi in cimette in acqua salata per 5 minuti, scolare e tenere l'acqua. 2. In una padella, sciogliere le acciughe in olio con aglio e peperoncino. 3. Aggiungere i broccoli e schiacciare con una forchetta. 4. Cuocere la pasta nell'acqua dei broccoli. 5. Scolare e mantecare in padella con i broccoli. 6. Servire con un filo d'olio a crudo.",
        },
        {
            "RICETTA": "Zuppa di farro con verdure",
            "INGREDIENTI": "250g di farro perlato, 2 carote, 2 zucchine, 1 porro, 200g di cavolo nero, 400g di pomodori pelati, brodo vegetale, rosmarino, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Soffriggere il porro affettato in poco olio. 2. Aggiungere carote e zucchine a dadini. 3. Unire il farro, i pelati e il brodo. 4. Aggiungere il cavolo nero e il rosmarino. 5. Cuocere per 30 minuti fino a che il farro è tenero. 6. Aggiustare di sale e pepe, servire con olio a crudo.",
        },
        {
            "RICETTA": "Pasta di legumi con pesto di rucola e noci",
            "INGREDIENTI": "300g di pasta di lenticchie rosse, 100g di rucola, 30g di noci, 30g di parmigiano, 1 spicchio d'aglio, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Frullare rucola, noci, parmigiano, aglio e olio fino a ottenere un pesto cremoso. 2. Cuocere la pasta di legumi al dente seguendo le indicazioni. 3. Scolare tenendo da parte un po' di acqua di cottura. 4. Condire con il pesto, aggiungendo acqua di cottura per amalgamare. 5. Servire con noci tritate e parmigiano. 6. Aggiungere pepe nero a piacere.",
        },
        {
            "RICETTA": "Minestrone di legumi misti",
            "INGREDIENTI": "200g di legumi misti secchi (ceci, fagioli, lenticchie), 2 carote, 2 patate, 1 zucchina, 1 cipolla, 1 sedano, 200g di passata di pomodoro, rosmarino, salvia, olio, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "proteica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Ammollare i legumi per 12 ore, scolare e risciacquare. 2. Soffriggere cipolla, sedano e carote tritati. 3. Aggiungere patate e zucchina a dadini. 4. Unire i legumi, la passata, le erbe aromatiche e coprire con acqua. 5. Cuocere a fuoco dolce per 1 ora e mezza. 6. Servire con olio a crudo e pane integrale tostato.",
        },
        {
            "RICETTA": "Risotto integrale con zucca e rosmarino",
            "INGREDIENTI": "300g di riso integrale, 500g di zucca, 1 cipolla, brodo vegetale, rosmarino, 30g di parmigiano, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare la zucca a cubetti piccoli. 2. Soffriggere la cipolla tritata in olio. 3. Tostare il riso per 2 minuti, poi aggiungere la zucca. 4. Aggiungere brodo caldo un mestolo alla volta, mescolando. 5. Cuocere per 40-45 minuti fino a che il riso è tenero. 6. Mantecare con parmigiano, rosmarino tritato e pepe nero.",
        },
        {
            "RICETTA": "Crema di ceci e zenzero",
            "INGREDIENTI": "400g di ceci cotti, 1 cipolla, 1 carota, 1 cucchiaino di zenzero fresco grattugiato, 1 cucchiaino di paprika, brodo vegetale, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Soffriggere cipolla e carota tritata in poco olio. 2. Aggiungere zenzero e paprika, mescolare per 1 minuto. 3. Unire i ceci e il brodo. 4. Cuocere per 15 minuti. 5. Frullare il tutto fino a ottenere una crema liscia. 6. Servire con crostini integrali, un filo d'olio e paprika.",
        },
        {
            "RICETTA": "Pasta integrale al sugo di lenticchie",
            "INGREDIENTI": "320g di fusilli integrali, 200g di lenticchie verdi, 400g di passata di pomodoro, 1 carota, 1 sedano, 1 cipolla, alloro, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Soffriggere cipolla, carota e sedano tritati in olio. 2. Aggiungere le lenticchie e coprire con acqua. 3. Cuocere 20 minuti con alloro. 4. Aggiungere la passata e cuocere altri 15 minuti. 5. Cuocere la pasta al dente e condire con il sugo. 6. Servire con un filo d'olio a crudo.",
        },
    ],
    "SECONDI": [
        {
            "RICETTA": "Salmone al forno con crosta di erbe",
            "INGREDIENTI": "4 filetti di salmone (150g ciascuno), pangrattato integrale, prezzemolo, timo, aglio, limone, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Mescolare pangrattato con erbe tritate, aglio, scorza di limone e olio. 2. Adagiare i filetti in teglia con carta forno. 3. Distribuire la crosta di erbe sui filetti premendo leggermente. 4. Cuocere in forno a 200°C per 15-18 minuti. 5. Il salmone deve essere rosato all'interno. 6. Servire con limone e insalata mista.",
        },
        {
            "RICETTA": "Petto di pollo alla griglia con limone e origano",
            "INGREDIENTI": "4 petti di pollo (150g ciascuno), succo di 2 limoni, origano, aglio, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Marinare i petti con limone, origano, aglio tritato, olio, sale e pepe per almeno 30 minuti. 2. Scaldare una griglia o padella a fuoco medio-alto. 3. Cuocere 5-6 minuti per lato senza muovere. 4. Far riposare 5 minuti prima di tagliare. 5. Servire affettato con verdure grigliate. 6. Ottimo anche freddo per insalate.",
        },
        {
            "RICETTA": "Polpette di tacchino e zucchine",
            "INGREDIENTI": "400g di macinato di tacchino, 2 zucchine, 1 uovo, 40g di pangrattato integrale, menta, aglio, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Grattugiare le zucchine e strizzarle bene. 2. Mescolare tacchino, zucchine, uovo, pangrattato, menta tritata, aglio e sale. 3. Formare polpette della dimensione di una noce. 4. Cuocere in forno a 190°C per 20-25 minuti. 5. Girarle a metà cottura. 6. Servire con insalata o yogurt greco alle erbe.",
        },
        {
            "RICETTA": "Sgombro al forno con pomodorini e olive",
            "INGREDIENTI": "4 filetti di sgombro, 300g di pomodorini, 80g di olive taggiasche, capperi, aglio, origano, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare i pomodorini a metà, mescolare con olive, capperi, aglio e origano. 2. Disporre i filetti di sgombro in teglia con carta forno. 3. Distribuire il condimento sui filetti. 4. Condire con olio e sale. 5. Cuocere in forno a 200°C per 15-20 minuti. 6. Servire con pane integrale.",
        },
        {
            "RICETTA": "Frittata di verdure al forno",
            "INGREDIENTI": "6 uova, 200g di spinaci, 1 zucchina, 1 peperone, 1 cipolla, 30g di parmigiano, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Saltare le verdure tagliate a dadini in padella con olio per 8 minuti. 2. Sbattere le uova con parmigiano, sale e pepe. 3. Unire le verdure alle uova. 4. Versare in una teglia con carta forno. 5. Cuocere in forno a 180°C per 25 minuti. 6. Servire tiepida o fredda, ottima anche per il pranzo al sacco.",
        },
        {
            "RICETTA": "Merluzzo con crema di cannellini",
            "INGREDIENTI": "4 filetti di merluzzo (150g ciascuno), 400g di fagioli cannellini cotti, 1 spicchio d'aglio, rosmarino, limone, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Frullare i cannellini con aglio, rosmarino, un filo d'olio e succo di limone fino a crema. 2. Scaldare la crema in padella a fuoco dolce. 3. In un'altra padella, cuocere i filetti di merluzzo con olio per 4 minuti per lato. 4. Disporre la crema nel piatto, adagiare il pesce sopra. 5. Condire con olio, pepe e scorza di limone. 6. Servire con verdure al vapore.",
        },
        {
            "RICETTA": "Straccetti di manzo con rucola e grana",
            "INGREDIENTI": "400g di fesa di manzo a fettine sottili, 100g di rucola, 50g di grana a scaglie, succo di limone, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare la carne a strisce sottili. 2. Scaldare una padella antiaderente a fuoco vivace con poco olio. 3. Cuocere la carne per 2 minuti, mantenendola rosata. 4. Trasferire in un piatto con rucola fresca. 5. Condire con limone, olio, sale e pepe. 6. Completare con scaglie di grana.",
        },
    ],
    "PIATTI UNICI": [
        {
            "RICETTA": "Bowl di quinoa con ceci e verdure arrosto",
            "INGREDIENTI": "200g di quinoa, 400g di ceci cotti, 2 zucchine, 1 peperone, 1 melanzana, 1 cipolla rossa, olio, paprika, cumino, limone, tahina, sale",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Cuocere la quinoa in acqua salata per 15 minuti, scolare. 2. Tagliare le verdure a pezzi, condire con olio, paprika e cumino. 3. Arrostire in forno a 200°C per 25 minuti. 4. Scaldare i ceci in padella con un pizzico di cumino. 5. Comporre la bowl con quinoa, verdure arrosto e ceci. 6. Condire con salsa di tahina e limone.",
        },
        {
            "RICETTA": "Insalata di farro con tonno e fagioli",
            "INGREDIENTI": "250g di farro perlato, 200g di tonno al naturale, 200g di fagioli cannellini, 100g di pomodorini, 1 cetriolo, cipolla rossa, basilico, olio, sale, pepe",
            "STAGIONALITA": "Estate",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Cuocere il farro in acqua salata per 25 minuti, scolare e raffreddare. 2. Scolare tonno e fagioli. 3. Tagliare pomodorini, cetriolo e cipolla rossa. 4. Mescolare tutti gli ingredienti in una ciotola capiente. 5. Condire con olio, sale, pepe e basilico fresco. 6. Servire fresca, perfetta anche da portare via.",
        },
        {
            "RICETTA": "Curry di ceci e spinaci con riso basmati",
            "INGREDIENTI": "400g di ceci cotti, 300g di spinaci freschi, 200g di riso basmati, 200ml di latte di cocco, 1 cipolla, 2 pomodori, aglio, zenzero, curry in polvere, curcuma, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Soffriggere cipolla, aglio e zenzero tritati. 2. Aggiungere curry e curcuma, mescolare per 1 minuto. 3. Unire i pomodori a pezzi e cuocere 5 minuti. 4. Aggiungere ceci, latte di cocco e cuocere 15 minuti. 5. Unire gli spinaci e cuocere fino ad appassimento. 6. Servire con riso basmati cotto al vapore.",
        },
        {
            "RICETTA": "Tortilla di patate e cipolle",
            "INGREDIENTI": "6 uova, 500g di patate, 2 cipolle, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": "glucidica",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare le patate a fettine sottili e le cipolle a rondelle. 2. Friggere a fuoco medio in olio per 15 minuti mescolando. 3. Sbattere le uova con sale e pepe. 4. Unire patate e cipolle alle uova. 5. Versare in padella antiaderente e cuocere a fuoco dolce per 8 minuti. 6. Capovolgere con un piatto e completare la cottura per 5 minuti.",
        },
        {
            "RICETTA": "Piadina integrale con hummus e verdure grigliate",
            "INGREDIENTI": "4 piadine integrali, 200g di hummus, 2 zucchine, 1 peperone, 1 melanzana, 100g di feta, rucola, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare le verdure a fette e grigliarle su piastra ben calda. 2. Salare e condire con olio le verdure grigliate. 3. Scaldare le piadine sulla piastra per 1 minuto per lato. 4. Spalmare l'hummus sulla piadina. 5. Farcire con verdure, feta sbriciolata e rucola. 6. Chiudere a metà e servire.",
        },
        {
            "RICETTA": "Zuppa di orzo con funghi e fagioli",
            "INGREDIENTI": "200g di orzo perlato, 200g di funghi champignon, 200g di fagioli borlotti cotti, 1 carota, 1 sedano, 1 cipolla, timo, brodo vegetale, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Soffriggere cipolla, carota e sedano tritati in olio. 2. Aggiungere i funghi affettati e cuocere 5 minuti. 3. Unire l'orzo, i fagioli e il brodo. 4. Aggiungere il timo e cuocere per 30 minuti. 5. Aggiustare di sale e pepe. 6. Servire con un filo d'olio a crudo e pane integrale.",
        },
        {
            "RICETTA": "Insalatona proteica con uova e avocado",
            "INGREDIENTI": "4 uova, 2 avocado, 200g di pomodorini, 1 cetriolo, 100g di mais, 100g di fagioli neri, lattuga mista, semi di girasole, limone, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "proteica",
            "FONTE 2": "fibra",
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Cuocere le uova sode per 9 minuti, raffreddare e tagliare a spicchi. 2. Tagliare avocado, pomodorini e cetriolo. 3. Scolare mais e fagioli neri. 4. Comporre l'insalata in un piatto con la lattuga alla base. 5. Disporre tutti gli ingredienti sopra. 6. Condire con olio, limone, sale e semi di girasole.",
        },
        {
            "RICETTA": "Riso venere con gamberi e zucchine",
            "INGREDIENTI": "300g di riso venere, 300g di gamberi sgusciati, 2 zucchine, 1 spicchio d'aglio, prezzemolo, limone, olio, sale, pepe",
            "STAGIONALITA": "Estate",
            "FONTE": "glucidica",
            "FONTE 2": "proteica",
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Cuocere il riso venere in acqua salata per 40 minuti, scolare e raffreddare. 2. Tagliare le zucchine a dadini e saltarle in padella con aglio. 3. Nella stessa padella, cuocere i gamberi per 2-3 minuti per lato. 4. Mescolare riso, zucchine e gamberi. 5. Condire con olio, limone, prezzemolo tritato, sale e pepe. 6. Servire tiepido o freddo.",
        },
    ],
    "CONTORNI": [
        {
            "RICETTA": "Caponata siciliana leggera",
            "INGREDIENTI": "3 melanzane, 2 zucchine, 1 peperone, 200g di pomodorini, 1 cipolla, sedano, olive verdi, capperi, aceto di mele, basilico, olio, sale",
            "STAGIONALITA": "Estate",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "media",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare melanzane e zucchine a cubetti, salare e far scolare per 20 minuti. 2. Cuocere le verdure separatamente in padella con poco olio. 3. Soffriggere cipolla e sedano. 4. Unire tutte le verdure, pomodorini, olive e capperi. 5. Aggiungere aceto di mele e cuocere 10 minuti. 6. Servire fredda con basilico fresco.",
        },
        {
            "RICETTA": "Insalata di finocchi e arance",
            "INGREDIENTI": "3 finocchi, 2 arance, olive nere, olio, sale, pepe nero",
            "STAGIONALITA": "Inverno",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare i finocchi a fettine sottili con la mandolina. 2. Pelare le arance a vivo e tagliare a rondelle. 3. Disporre finocchi e arance in un piatto. 4. Aggiungere olive nere. 5. Condire con olio, sale e pepe nero. 6. Far riposare 10 minuti prima di servire.",
        },
        {
            "RICETTA": "Verdure al forno miste",
            "INGREDIENTI": "2 zucchine, 2 peperoni, 1 melanzana, 1 cipolla rossa, 300g di pomodorini, aglio, timo, origano, olio, sale",
            "STAGIONALITA": "All",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Tagliare tutte le verdure a pezzi regolari. 2. Condire con olio, aglio, timo, origano e sale. 3. Disporre in teglia in un singolo strato. 4. Cuocere in forno a 200°C per 30-35 minuti. 5. Mescolare a metà cottura. 6. Servire calde o a temperatura ambiente.",
        },
        {
            "RICETTA": "Cavolo nero saltato con aglio e peperoncino",
            "INGREDIENTI": "500g di cavolo nero, 2 spicchi d'aglio, peperoncino, olio, sale",
            "STAGIONALITA": "Inverno",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Eliminare le coste dure del cavolo nero e lavare bene le foglie. 2. Sbollentare per 3 minuti in acqua salata, scolare. 3. In padella, scaldare olio con aglio e peperoncino. 4. Aggiungere il cavolo nero e saltare per 5 minuti. 5. Salare e servire con un filo d'olio a crudo. 6. Ottimo anche come condimento per bruschette.",
        },
        {
            "RICETTA": "Purè di cavolfiore",
            "INGREDIENTI": "1 cavolfiore grande, 30g di parmigiano, noce moscata, olio, sale, pepe",
            "STAGIONALITA": "Inverno",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Dividere il cavolfiore in cimette e cuocere al vapore per 15 minuti. 2. Frullare con un filo d'olio fino a ottenere un purè liscio. 3. Aggiungere parmigiano e noce moscata. 4. Aggiustare di sale e pepe. 5. Servire caldo come alternativa leggera al purè di patate. 6. Perfetto come accompagnamento a carne e pesce.",
        },
        {
            "RICETTA": "Barbabietole al forno con yogurt e noci",
            "INGREDIENTI": "4 barbabietole, 150g di yogurt greco, 30g di noci, aceto balsamico, erba cipollina, olio, sale, pepe",
            "STAGIONALITA": "All",
            "FONTE": "fibra",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Avvolgere le barbabietole in alluminio e cuocere in forno a 200°C per 1 ora. 2. Lasciar raffreddare e pelare. 3. Tagliare a spicchi e condire con olio, aceto balsamico e sale. 4. Disporre nel piatto con cucchiaiate di yogurt greco. 5. Completare con noci spezzettate e erba cipollina. 6. Servire tiepida o a temperatura ambiente.",
        },
        {
            "RICETTA": "Edamame saltati con sesamo",
            "INGREDIENTI": "400g di edamame sgusciati, 1 spicchio d'aglio, salsa di soia, olio di sesamo, semi di sesamo, peperoncino",
            "STAGIONALITA": "All",
            "FONTE": "proteica",
            "FONTE 2": None,
            "DIFFICOLTA": "facile",
            "LIBRO_PAG": None,
            "PROCEDIMENTO": "1. Cuocere gli edamame in acqua bollente per 3 minuti, scolare. 2. Scaldare olio di sesamo in padella con aglio. 3. Saltare gli edamame per 3 minuti a fuoco vivace. 4. Aggiungere salsa di soia e peperoncino. 5. Completare con semi di sesamo tostati. 6. Servire come snack proteico o contorno.",
        },
    ],
}


def main():
    wb = Workbook()
    wb.remove(wb.active)

    for category in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
        ws = wb.create_sheet(category)
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for recipe in RECIPES[category]:
            ws.append([recipe[col] for col in COLUMNS])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"File salvato: {OUTPUT}")
    for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
        print(f"  {cat}: {len(RECIPES[cat])} ricette")
    print(f"  Totale: {sum(len(v) for v in RECIPES.values())} ricette")


if __name__ == "__main__":
    main()
