"""Genera il file dati/ingredienti.xlsx con tutti gli ingredienti unici
estratti dalle ricette, classificati per categoria alimentare e con
valori nutrizionali per 100g.

Output: dati/ingredienti.xlsx (9 sheet, uno per categoria)
"""

import sys
sys.path.insert(0, ".")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.excel import RecipeRepository
from src.list_generator import IngredientParser
from src.model import UnitNormalizer

# ---------------------------------------------------------------------------
# Valori nutrizionali per 100g: {nome: (kcal, proteine, carbo, grassi, fibre)}
# ---------------------------------------------------------------------------

NUTRITION_DB: dict[str, tuple[float, float, float, float, float]] = {
    # --- Verdure ---
    "zucchine": (17, 1.2, 3.1, 0.3, 1.0),
    "broccoli": (34, 2.8, 7.0, 0.4, 2.6),
    "melanzane": (25, 1.0, 6.0, 0.2, 3.0),
    "peperoni": (31, 1.0, 6.0, 0.3, 2.1),
    "asparagi": (20, 2.2, 3.9, 0.1, 2.1),
    "carciofi": (47, 3.3, 10.5, 0.2, 5.4),
    "piselli": (81, 5.4, 14.5, 0.4, 5.1),
    "spinaci": (23, 2.9, 3.6, 0.4, 2.2),
    "funghi champignon": (22, 3.1, 3.3, 0.3, 1.0),
    "funghi": (22, 3.1, 3.3, 0.3, 1.0),
    "funghi porcini": (26, 3.3, 3.7, 0.3, 1.5),
    "funghi misti": (24, 3.0, 3.5, 0.3, 1.2),
    "cavolfiore": (25, 1.9, 5.0, 0.3, 2.0),
    "pomodorini": (18, 0.9, 3.9, 0.2, 1.2),
    "pomodori": (18, 0.9, 3.9, 0.2, 1.2),
    "pomodori pelati": (18, 0.9, 3.9, 0.2, 1.2),
    "passata di pomodoro": (24, 1.0, 4.5, 0.2, 1.5),
    "pomodoro": (18, 0.9, 3.9, 0.2, 1.2),
    "fagiolini": (31, 1.8, 7.0, 0.1, 3.4),
    "zucca": (26, 1.0, 6.5, 0.1, 0.5),
    "radicchio": (23, 1.4, 4.5, 0.3, 0.9),
    "cicoria": (23, 1.7, 4.7, 0.3, 4.0),
    "finocchi": (31, 1.2, 7.3, 0.2, 3.1),
    "bietole": (19, 1.8, 3.7, 0.2, 1.6),
    "cime di rapa": (22, 2.9, 2.9, 0.3, 2.7),
    "verza": (25, 1.3, 5.8, 0.1, 2.5),
    "cavolo nero": (35, 3.3, 4.4, 0.7, 4.1),
    "sedano rapa": (42, 1.5, 9.2, 0.3, 1.8),
    "rape rosse": (43, 1.6, 9.6, 0.2, 2.8),
    "barbabietola": (43, 1.6, 9.6, 0.2, 2.8),
    "carote": (41, 0.9, 9.6, 0.2, 2.8),
    "carota": (41, 0.9, 9.6, 0.2, 2.8),
    "porri": (61, 1.5, 14.2, 0.3, 1.8),
    "porro": (61, 1.5, 14.2, 0.3, 1.8),
    "cipolla": (40, 1.1, 9.3, 0.1, 1.7),
    "cipolla rossa": (40, 1.1, 9.3, 0.1, 1.7),
    "cipolle": (40, 1.1, 9.3, 0.1, 1.7),
    "cipollotto": (32, 1.8, 7.3, 0.2, 2.6),
    "topinambur": (73, 2.0, 17.4, 0.0, 1.6),
    "sedano": (16, 0.7, 3.0, 0.2, 1.6),
    "cetriolo": (15, 0.7, 3.6, 0.1, 0.5),
    "cetrioli": (15, 0.7, 3.6, 0.1, 0.5),
    "lattuga": (15, 1.4, 2.9, 0.2, 1.3),
    "rucola": (25, 2.6, 3.7, 0.7, 1.6),
    "cavolo cappuccio": (25, 1.3, 5.8, 0.1, 2.5),
    "patate": (77, 2.0, 17.5, 0.1, 2.2),
    "patata": (77, 2.0, 17.5, 0.1, 2.2),
    "patate dolci": (86, 1.6, 20.1, 0.1, 3.0),
    "mais": (86, 3.3, 19.0, 1.4, 2.7),
    "insalata mista": (17, 1.3, 3.0, 0.2, 1.5),
    "scarola": (17, 1.3, 3.1, 0.2, 3.1),
    "catalogna": (23, 1.7, 4.7, 0.3, 4.0),
    "friarielli": (22, 2.9, 2.9, 0.3, 2.7),
    "broccoletti": (34, 2.8, 7.0, 0.4, 2.6),
    "cavolini di bruxelles": (43, 3.4, 9.0, 0.3, 3.8),
    "cipolline borettane": (40, 1.1, 9.3, 0.1, 1.7),
    # --- Carboidrati ---
    "pasta": (352, 12.5, 71.2, 1.5, 2.5),
    "spaghetti": (352, 12.5, 71.2, 1.5, 2.5),
    "penne": (352, 12.5, 71.2, 1.5, 2.5),
    "fusilli": (352, 12.5, 71.2, 1.5, 2.5),
    "rigatoni": (352, 12.5, 71.2, 1.5, 2.5),
    "farfalle": (352, 12.5, 71.2, 1.5, 2.5),
    "orecchiette": (352, 12.5, 71.2, 1.5, 2.5),
    "linguine": (352, 12.5, 71.2, 1.5, 2.5),
    "tagliatelle": (352, 12.5, 71.2, 1.5, 2.5),
    "paccheri": (352, 12.5, 71.2, 1.5, 2.5),
    "conchiglie": (352, 12.5, 71.2, 1.5, 2.5),
    "casarecce": (352, 12.5, 71.2, 1.5, 2.5),
    "trofie": (352, 12.5, 71.2, 1.5, 2.5),
    "bucatini": (352, 12.5, 71.2, 1.5, 2.5),
    "mezze maniche": (352, 12.5, 71.2, 1.5, 2.5),
    "sedanini": (352, 12.5, 71.2, 1.5, 2.5),
    "ditalini": (352, 12.5, 71.2, 1.5, 2.5),
    "tubetti": (352, 12.5, 71.2, 1.5, 2.5),
    "pasta mista corta": (352, 12.5, 71.2, 1.5, 2.5),
    "pasta corta": (352, 12.5, 71.2, 1.5, 2.5),
    "spaghetti spezzati": (352, 12.5, 71.2, 1.5, 2.5),
    "penne integrali": (338, 13.4, 66.2, 2.5, 7.0),
    "fusilli integrali": (338, 13.4, 66.2, 2.5, 7.0),
    "spaghetti integrali": (338, 13.4, 66.2, 2.5, 7.0),
    "gnocchi": (133, 3.0, 30.0, 0.5, 1.5),
    "riso": (360, 7.0, 79.0, 0.6, 1.3),
    "riso carnaroli": (360, 7.0, 79.0, 0.6, 1.3),
    "riso basmati": (350, 7.1, 78.0, 0.6, 0.4),
    "riso integrale": (362, 7.5, 76.2, 2.7, 3.5),
    "riso venere": (356, 8.5, 72.0, 2.8, 3.0),
    "farro": (338, 14.6, 67.1, 2.5, 6.8),
    "orzo perlato": (352, 9.9, 77.7, 1.2, 5.0),
    "quinoa": (368, 14.1, 64.2, 6.1, 7.0),
    "cous-cous": (376, 12.8, 77.4, 0.6, 5.0),
    "bulgur": (342, 12.3, 75.9, 1.3, 12.5),
    "miglio": (378, 11.0, 72.9, 4.2, 3.5),
    "polenta": (70, 1.6, 15.7, 0.3, 1.0),
    "pane": (265, 8.8, 49.0, 3.3, 2.7),
    "pane raffermo": (265, 8.8, 49.0, 3.3, 2.7),
    "pangrattato": (395, 11.0, 72.0, 5.5, 3.5),
    "tortillas integrali": (310, 9.0, 49.0, 8.0, 5.0),
    "pita integrali": (275, 10.0, 55.0, 2.0, 6.0),
    "brodo vegetale": (4, 0.2, 0.6, 0.1, 0.0),
    # --- Pesce ---
    "salmone": (208, 20.4, 0.0, 13.4, 0.0),
    "salmone affumicato": (176, 24.5, 0.0, 8.4, 0.0),
    "merluzzo": (82, 17.8, 0.0, 0.7, 0.0),
    "orata": (100, 20.7, 0.0, 1.8, 0.0),
    "branzino": (97, 18.4, 0.0, 2.5, 0.0),
    "tonno fresco": (144, 23.3, 0.0, 4.9, 0.0),
    "tonno": (144, 23.3, 0.0, 4.9, 0.0),
    "sgombro": (205, 18.6, 0.0, 13.9, 0.0),
    "pesce spada": (121, 19.8, 0.0, 4.4, 0.0),
    "sogliola": (86, 17.5, 0.0, 1.4, 0.0),
    "trota": (141, 19.7, 0.0, 6.6, 0.0),
    "sardine": (208, 24.6, 0.0, 11.5, 0.0),
    "gamberi": (85, 20.1, 0.0, 0.5, 0.0),
    "calamari": (92, 15.6, 3.1, 1.4, 0.0),
    "cozze": (86, 11.9, 3.7, 2.2, 0.0),
    "vongole": (72, 12.8, 2.6, 1.0, 0.0),
    "acciughe": (131, 20.4, 0.0, 4.8, 0.0),
    "polpo": (82, 14.9, 2.2, 1.0, 0.0),
    "scampi": (90, 18.0, 0.0, 1.5, 0.0),
    # --- Carne ---
    "petto di pollo": (165, 31.0, 0.0, 3.6, 0.0),
    "pollo": (165, 31.0, 0.0, 3.6, 0.0),
    "pollo grigliato": (165, 31.0, 0.0, 3.6, 0.0),
    "sovracosce di pollo": (177, 25.1, 0.0, 8.1, 0.0),
    "fettine di tacchino": (135, 30.0, 0.0, 1.5, 0.0),
    "tacchino": (135, 30.0, 0.0, 1.5, 0.0),
    "macinato di tacchino": (148, 27.0, 0.0, 4.0, 0.0),
    "fettine di vitello": (172, 28.0, 0.0, 6.0, 0.0),
    "vitello": (172, 28.0, 0.0, 6.0, 0.0),
    "macinato di manzo": (254, 17.2, 0.0, 20.0, 0.0),
    "manzo": (250, 26.0, 0.0, 15.0, 0.0),
    "tagliata di manzo": (250, 26.0, 0.0, 15.0, 0.0),
    "straccetti di manzo": (250, 26.0, 0.0, 15.0, 0.0),
    "scaloppine": (172, 28.0, 0.0, 6.0, 0.0),
    "hamburger di manzo": (254, 17.2, 0.0, 20.0, 0.0),
    "lonza di maiale": (143, 26.0, 0.0, 3.5, 0.0),
    "coniglio": (136, 20.0, 0.0, 6.0, 0.0),
    "salsiccia": (297, 13.6, 1.0, 26.3, 0.0),
    "pancetta": (393, 14.4, 0.0, 37.0, 0.0),
    "agnello": (282, 16.6, 0.0, 23.4, 0.0),
    # --- Legumi ---
    "ceci": (164, 8.9, 27.4, 2.6, 7.6),
    "lenticchie": (116, 9.0, 20.1, 0.4, 7.9),
    "lenticchie rosse": (116, 9.0, 20.1, 0.4, 7.9),
    "fagioli cannellini": (91, 6.2, 16.5, 0.3, 6.3),
    "fagioli borlotti": (124, 9.0, 22.5, 0.5, 7.0),
    "fagioli neri": (132, 8.9, 23.7, 0.5, 8.7),
    "fagioli": (91, 6.2, 16.5, 0.3, 6.3),
    "fave": (88, 7.6, 17.6, 0.7, 5.4),
    "edamame": (122, 11.9, 8.9, 5.2, 5.2),
    "piselli secchi": (118, 8.3, 21.1, 0.4, 8.3),
    # --- Frutta ---
    "limone": (29, 1.1, 9.3, 0.3, 2.8),
    "avocado": (160, 2.0, 8.5, 14.7, 6.7),
    "arancia": (47, 0.9, 11.8, 0.1, 2.4),
    "arance": (47, 0.9, 11.8, 0.1, 2.4),
    "olive": (145, 1.0, 3.8, 15.3, 3.3),
    "mela": (52, 0.3, 13.8, 0.2, 2.4),
    "mele": (52, 0.3, 13.8, 0.2, 2.4),
    "melograno": (83, 1.7, 18.7, 1.2, 4.0),
    "prugne": (46, 0.7, 11.4, 0.3, 1.4),
    "pomodori secchi": (258, 14.1, 43.5, 3.0, 12.3),
    # --- Latticini/Uova ---
    "uova": (155, 13.0, 1.1, 11.0, 0.0),
    "uovo": (155, 13.0, 1.1, 11.0, 0.0),
    "uova sode": (155, 13.0, 1.1, 11.0, 0.0),
    "parmigiano": (392, 35.8, 3.2, 25.8, 0.0),
    "pecorino": (387, 25.5, 3.6, 32.0, 0.0),
    "ricotta": (174, 11.3, 3.0, 13.0, 0.0),
    "feta": (264, 14.2, 4.1, 21.3, 0.0),
    "mozzarella": (280, 22.2, 2.2, 17.1, 0.0),
    "gorgonzola": (353, 19.4, 0.0, 30.6, 0.0),
    "caprino": (280, 18.5, 1.0, 22.0, 0.0),
    "burrata": (290, 16.0, 1.5, 24.0, 0.0),
    "scamorza": (300, 24.0, 1.0, 22.0, 0.0),
    "yogurt greco": (97, 9.0, 3.6, 5.0, 0.0),
    "besciamella": (122, 4.0, 9.0, 8.0, 0.0),
    "burro": (717, 0.9, 0.1, 81.1, 0.0),
    "latte di cocco": (197, 2.0, 2.8, 21.3, 0.0),
    # --- Condimenti/Spezie ---
    "olio": (884, 0.0, 0.0, 100.0, 0.0),
    "olio di sesamo": (884, 0.0, 0.0, 100.0, 0.0),
    "aceto balsamico": (88, 0.5, 17.0, 0.0, 0.0),
    "aceto di mele": (22, 0.0, 1.0, 0.0, 0.0),
    "sale": (0, 0.0, 0.0, 0.0, 0.0),
    "pepe": (251, 10.4, 64.8, 3.3, 25.3),
    "peperoncino": (40, 1.9, 8.8, 0.4, 1.5),
    "aglio": (149, 6.4, 33.1, 0.5, 2.1),
    "basilico": (23, 3.2, 2.7, 0.6, 1.6),
    "prezzemolo": (36, 3.0, 6.3, 0.8, 3.3),
    "menta": (44, 3.3, 8.4, 0.7, 6.8),
    "rosmarino": (131, 3.3, 20.7, 5.9, 14.1),
    "timo": (101, 5.6, 24.5, 1.7, 14.0),
    "origano": (265, 9.0, 68.9, 4.3, 42.5),
    "erba cipollina": (30, 3.3, 4.4, 0.7, 2.5),
    "salvia": (315, 10.6, 60.7, 12.7, 40.3),
    "maggiorana": (271, 12.7, 60.6, 7.0, 40.3),
    "cumino": (375, 17.8, 44.2, 22.3, 10.5),
    "paprika": (282, 14.1, 53.9, 13.0, 34.9),
    "paprika affumicata": (282, 14.1, 53.9, 13.0, 34.9),
    "curcuma": (312, 9.7, 67.1, 3.3, 22.7),
    "curry": (325, 14.3, 55.8, 14.0, 33.2),
    "zafferano": (310, 11.4, 65.4, 5.9, 3.9),
    "noce moscata": (525, 5.8, 49.3, 36.3, 20.8),
    "salsa di soia": (53, 8.1, 4.9, 0.6, 0.8),
    "zucchero": (400, 0.0, 100.0, 0.0, 0.0),
    "miele": (304, 0.3, 82.4, 0.0, 0.2),
    "senape": (66, 4.4, 5.8, 3.3, 3.0),
    "capperi": (23, 2.4, 4.9, 0.9, 3.2),
    "tahina": (595, 17.0, 21.2, 53.8, 9.3),
    "pesto": (418, 5.3, 6.0, 42.0, 2.0),
    "hummus": (166, 7.9, 14.3, 9.6, 6.0),
    "vino bianco": (82, 0.1, 2.6, 0.0, 0.0),
    "zenzero": (80, 1.8, 17.8, 0.8, 2.0),
    "mentuccia": (44, 3.3, 8.4, 0.7, 6.8),
    "limone confit": (29, 1.1, 9.3, 0.3, 2.8),
    # --- Frutta secca/Semi ---
    "noci": (654, 15.2, 13.7, 65.2, 6.7),
    "mandorle": (579, 21.2, 21.6, 49.9, 12.5),
    "pinoli": (673, 13.7, 13.1, 68.4, 3.7),
    "nocciole": (628, 15.0, 16.7, 60.8, 9.7),
    "anacardi": (553, 18.2, 30.2, 43.9, 3.3),
    "semi di zucca": (559, 30.2, 10.7, 49.1, 6.0),
    "semi di sesamo": (573, 17.7, 23.5, 49.7, 11.8),
    "semi di girasole": (584, 20.8, 20.0, 51.5, 8.6),
    # --- Tofu/Tempeh ---
    "tofu": (76, 8.1, 1.9, 4.8, 0.3),
    "tempeh": (192, 20.3, 7.6, 10.8, 0.0),
    # --- Extra ---
    "guanciale": (655, 7.0, 0.0, 69.6, 0.0),
    "prosciutto cotto": (215, 19.8, 0.9, 14.7, 0.0),
    "bresaola": (151, 33.1, 0.0, 2.6, 0.0),
    "lasagne": (352, 12.5, 71.2, 1.5, 2.5),
    "tonnarelli": (352, 12.5, 71.2, 1.5, 2.5),
    "tortellini": (300, 12.0, 42.0, 9.0, 2.0),
    "farina": (340, 11.0, 72.0, 1.0, 2.7),
    "frollini": (480, 6.5, 65.0, 20.0, 2.0),
    "alloro": (313, 7.6, 74.9, 8.4, 26.3),
    "aneto": (43, 3.5, 7.0, 1.1, 2.1),
    "lievito": (105, 8.4, 18.3, 1.9, 8.1),
    "lime": (30, 0.7, 10.5, 0.2, 2.8),
    "acqua": (0, 0.0, 0.0, 0.0, 0.0),
    "tabasco": (12, 0.5, 1.0, 0.8, 0.5),
    "marsala": (126, 0.1, 14.0, 0.0, 0.0),
    "ravanelli": (16, 0.7, 3.4, 0.1, 1.6),
    "squaquerone": (245, 13.0, 3.0, 20.0, 0.0),
    "stracchino": (300, 18.5, 0.0, 25.1, 0.0),
    "grana": (392, 35.8, 3.2, 25.8, 0.0),
    "panna": (337, 2.1, 2.8, 35.0, 0.0),
    "tuorli": (322, 16.0, 3.6, 27.0, 0.0),
    "confettura di lamponi": (250, 0.4, 60.0, 0.3, 2.5),
    "sugo": (50, 1.5, 7.0, 1.5, 1.5),
    "piadine integrali": (290, 8.5, 48.0, 7.0, 5.0),
    "verdure grigliate": (35, 1.5, 6.0, 0.8, 2.5),
}

# Default nutrizionali per categoria (kcal, prot, carbo, grassi, fibre)
CATEGORY_DEFAULTS: dict[str, tuple[float, float, float, float, float]] = {
    "Verdure": (25, 1.5, 5.0, 0.3, 2.0),
    "Carboidrati": (350, 10.0, 72.0, 1.5, 3.0),
    "Pesce": (100, 19.0, 0.0, 3.0, 0.0),
    "Carne": (180, 25.0, 0.0, 8.0, 0.0),
    "Legumi": (120, 8.0, 20.0, 0.5, 7.0),
    "Frutta": (50, 0.8, 12.0, 0.3, 2.5),
    "Latticini_Uova": (250, 18.0, 3.0, 18.0, 0.0),
    "Condimenti_Spezie": (150, 3.0, 20.0, 8.0, 3.0),
    "Altro": (100, 3.0, 15.0, 3.0, 1.5),
}

# ---------------------------------------------------------------------------
# Classificazione per categoria (keyword-based con priorità)
# ---------------------------------------------------------------------------

CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "Pesce": [
        "salmone", "merluzzo", "orata", "branzino", "tonno", "sgombro",
        "pesce spada", "sogliola", "trota", "sardine", "gamberi", "calamari",
        "cozze", "vongole", "acciughe", "acciuga", "aggiughe", "polpo",
        "scampi", "pesce",
    ],
    "Carne": [
        "pollo", "tacchino", "manzo", "vitello", "maiale", "lonza",
        "coniglio", "salsiccia", "pancetta", "agnello", "hamburger",
        "macinato", "scaloppine", "straccetti", "tagliata", "fettine",
        "sovracosce", "petto", "guanciale", "prosciutto", "bresaola",
        "besaola", "carne",
    ],
    "Legumi": [
        "ceci", "lenticchie", "fagioli", "cannellini", "fave", "edamame",
        "piselli secchi", "tofu", "tempeh",
    ],
    "Carboidrati": [
        "pasta", "spaghetti", "penne", "fusilli", "rigatoni", "farfalle",
        "orecchiette", "linguine", "tagliatelle", "paccheri", "conchiglie",
        "casarecce", "trofie", "bucatini", "mezze maniche", "sedanini",
        "ditalini", "tubetti", "gnocchi", "lasagne", "tonnarelli",
        "tortellini", "piadine", "frollini",
        "riso", "farro", "orzo", "quinoa", "cous-cous", "couscous", "bulgur",
        "miglio", "polenta", "pane", "pangrattato", "tortillas", "pita",
        "patate", "patata", "brodo", "farina",
    ],
    "Latticini_Uova": [
        "uova", "uovo", "tuorli", "mozzarella", "parmigiano", "pecorino",
        "ricotta", "feta", "gorgonzola", "caprino", "burrata", "scamorza",
        "yogurt", "besciamella", "burro", "latte", "formaggio", "grana",
        "squaquerello", "squaquerone", "panna", "stracchino",
    ],
    "Frutta": [
        "limone", "limoni", "avocado", "arancia", "arance", "olive", "mela",
        "mele", "melograno", "prugne", "pomodori secchi", "lime",
        "lamponi", "confettura",
    ],
    "Verdure": [
        "zucchine", "zucchina", "broccoli", "broccolo", "melanzane",
        "peperoni", "asparagi", "carciofi", "piselli", "spinaci", "funghi",
        "cavolfiore", "pomodori", "pomodoro", "pomodirini", "passata",
        "fagiolini", "zucca", "radicchio", "cicoria", "finocchi", "bietole",
        "cime di rapa", "verza", "cavolo", "sedano", "rape", "carote",
        "carota", "porri", "porro", "cipolla", "cipolle", "cipollotto",
        "cipollla",
        "topinambur", "cetriolo", "cetrioli", "lattuga", "rucola",
        "insalata", "mais", "scarola", "catalogna", "friarielli",
        "broccoletti", "cavolini", "ravanelli", "barbabietola", "bababietola",
        "verdure",
    ],
    "Condimenti_Spezie": [
        "olio", "aceto", "sale", "pepe", "peperoncino", "aglio",
        "basilico", "prezzemolo", "menta", "rosmarino", "timo",
        "origano", "erba cipollina", "salvia", "maggiorana",
        "cumino", "paprika", "paprica", "curcuma", "curry", "zafferano",
        "noce moscata", "salsa di soia", "zucchero", "miele",
        "senape", "capperi", "tahina", "pesto", "hummus",
        "vino", "marsala", "zenzero", "mentuccia", "scalogno",
        "noci", "mandorle", "pinoli", "nocciole", "anacardi",
        "semi di", "semi ", "alloro", "aneto", "lievito",
        "sugo", "tabasco", "spezie", "acqua",
    ],
}

# Ordine di priorità: keyword match viene valutato in questo ordine
CATEGORY_PRIORITY = [
    "Pesce", "Carne", "Legumi", "Carboidrati",
    "Latticini_Uova", "Frutta", "Verdure", "Condimenti_Spezie",
]


def classify_ingredient(name: str) -> str:
    """Classifica un ingrediente in una delle 9 categorie."""
    name_lower = name.lower()
    for category in CATEGORY_PRIORITY:
        for keyword in CATEGORY_KEYWORDS[category]:
            if keyword in name_lower:
                return category
    return "Altro"


def get_nutrition(name: str, category: str) -> tuple[float, float, float, float, float]:
    """Trova i valori nutrizionali per un ingrediente.

    Strategia: match esatto -> match parziale -> default per categoria.
    """
    name_lower = name.lower()

    # 1. Match esatto
    if name_lower in NUTRITION_DB:
        return NUTRITION_DB[name_lower]

    # 2. Match parziale (l'ingrediente contiene una chiave del DB o viceversa)
    best_match = None
    best_len = 0
    for db_name, values in NUTRITION_DB.items():
        if db_name in name_lower or name_lower in db_name:
            if len(db_name) > best_len:
                best_match = values
                best_len = len(db_name)
    if best_match:
        return best_match

    # 3. Default per categoria
    return CATEGORY_DEFAULTS.get(category, CATEGORY_DEFAULTS["Altro"])


# ---------------------------------------------------------------------------
# Estrazione ingredienti unici
# ---------------------------------------------------------------------------

# Ingredienti da ignorare (troppo generici o non alimentari)
IGNORE_INGREDIENTS = {
    "", "q.b.", "qb", "q.b", "a piacere", "quanto basta",
    # Frammenti di metodi di cottura (non ingredienti)
    "al forno", "in padella", "alla griglia", "al cartoccio",
    "al vapore", "gratinato", "al sesamo", "al marsala",
    "al tegamino con tartufo", "alla cacciatora", "alla diavola",
    "alla fiorentina", "alla marinara", "alla norma", "alla pizzaiola",
    "alla provenzale", "alla sorrentina", "alla piastra con verdure",
    "alle erbe aromatiche", "all'uccelletto", "con erbe aromatiche",
    "con ragù di verdure", "con verdure croccanti", "con verdure miste",
    "croccante con salsa agrodolce", "dolci al forno",
    "in purgatorio", "marinata alle erbe", "marinato alle erbe",
    "novelle al forno con erbe", "sode con salsa tonnata",
}


def extract_unique_ingredients(recipes_by_category: dict) -> set[str]:
    """Estrae tutti i nomi ingrediente unici dalle ricette."""
    parser = IngredientParser(UnitNormalizer())
    unique_names: set[str] = set()

    for category_recipes in recipes_by_category.values():
        for recipe in category_recipes:
            if not isinstance(recipe.ingredients, str):
                continue
            ingredients = parser.parse(recipe.ingredients)
            for ing in ingredients:
                name = ing.name.strip()
                if name and name not in IGNORE_INGREDIENTS and len(name) > 1:
                    unique_names.add(name)

    return unique_names


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

SHEET_COLORS = {
    "Verdure": "92D050",
    "Carboidrati": "FFC000",
    "Pesce": "5B9BD5",
    "Carne": "FF6B6B",
    "Legumi": "A9D18E",
    "Frutta": "F4B183",
    "Latticini_Uova": "FFFFAA",
    "Condimenti_Spezie": "D5A6E6",
    "Altro": "D9D9D9",
}

SHEET_DISPLAY_NAMES = {
    "Verdure": "Verdure",
    "Carboidrati": "Carboidrati",
    "Pesce": "Pesce",
    "Carne": "Carne",
    "Legumi": "Legumi",
    "Frutta": "Frutta",
    "Latticini_Uova": "Latticini e Uova",
    "Condimenti_Spezie": "Condimenti e Spezie",
    "Altro": "Altro",
}

HEADERS = ["Ingrediente", "Kcal (100g)", "Proteine (g)", "Carboidrati (g)", "Grassi (g)", "Fibre (g)"]


def write_ingredienti_excel(
    classified: dict[str, list[tuple[str, tuple[float, float, float, float, float]]]],
    output_path: str,
):
    """Scrive il file Excel con un sheet per categoria."""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_font = Font(size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    sheet_order = list(CATEGORY_PRIORITY) + ["Altro"]

    for category in sheet_order:
        items = classified.get(category, [])
        display_name = SHEET_DISPLAY_NAMES.get(category, category)
        ws = wb.create_sheet(title=display_name)

        fill_color = SHEET_COLORS.get(category, "D9D9D9")
        header_fill = PatternFill("solid", fgColor=fill_color)

        # Header
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # Dati ordinati alfabeticamente
        for row_idx, (name, nutrition) in enumerate(sorted(items, key=lambda x: x[0]), 2):
            kcal, prot, carb, fat, fiber = nutrition
            ws.cell(row=row_idx, column=1, value=name.capitalize()).font = data_font
            ws.cell(row=row_idx, column=1).alignment = align_left
            for col, val in enumerate([kcal, prot, carb, fat, fiber], 2):
                cell = ws.cell(row=row_idx, column=col, value=round(val, 1))
                cell.font = data_font
                cell.alignment = align_center

        # Auto-width
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    output_path = "dati/ingredienti.xlsx"

    # 1. Carica tutte le ricette
    print("Caricamento ricette...")
    repo = RecipeRepository("dati/ricette")
    recipes_by_category, _ = repo.load_recipes()
    total_recipes = sum(len(v) for v in recipes_by_category.values())
    print(f"  {total_recipes} ricette caricate")

    # 2. Estrai ingredienti unici
    print("\nEstrazione ingredienti unici...")
    unique_names = extract_unique_ingredients(recipes_by_category)
    print(f"  {len(unique_names)} ingredienti unici trovati")

    # 3. Classifica e assegna valori nutrizionali
    print("\nClassificazione per categoria...")
    classified: dict[str, list[tuple[str, tuple[float, float, float, float, float]]]] = {
        cat: [] for cat in list(CATEGORY_PRIORITY) + ["Altro"]
    }

    for name in unique_names:
        category = classify_ingredient(name)
        nutrition = get_nutrition(name, category)
        classified[category].append((name, nutrition))

    for cat in list(CATEGORY_PRIORITY) + ["Altro"]:
        display = SHEET_DISPLAY_NAMES.get(cat, cat)
        print(f"  {display}: {len(classified[cat])} ingredienti")

    # 4. Scrivi Excel
    print(f"\nScrittura {output_path}...")
    write_ingredienti_excel(classified, output_path)
    print(f"File creato: {output_path}")

    # 5. Riepilogo
    total_ingredients = sum(len(v) for v in classified.values())
    altro_count = len(classified["Altro"])
    print(f"\nRiepilogo:")
    print(f"  Totale ingredienti: {total_ingredients}")
    print(f"  Bucket 'Altro': {altro_count} ingredienti")
    if altro_count > 0:
        print(f"  Ingredienti in 'Altro': {', '.join(sorted(n for n, _ in classified['Altro']))}")


if __name__ == "__main__":
    main()
