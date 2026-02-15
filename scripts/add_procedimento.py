"""Script per aggiungere la colonna PROCEDIMENTO a Cucina ottimizzata.xlsx."""
import openpyxl

PROCEDIMENTI = {
    # === PRIMI ===
    "Risotto agli asparagi": (
        "1. Pulire gli asparagi, separare le punte e tagliare i gambi a rondelle. "
        "2. Soffriggere lo scalogno tritato in olio d'oliva. "
        "3. Aggiungere i gambi degli asparagi e tostare il riso per 2 minuti. "
        "4. Sfumare con il vino bianco. "
        "5. Aggiungere il brodo vegetale caldo un mestolo alla volta, mescolando spesso. "
        "6. A metà cottura aggiungere le punte degli asparagi. "
        "7. Mantecare con burro e parmigiano. Aggiustare di sale e pepe."
    ),
    "Piadina classica": (
        "1. Mescolare farina, strutto (o olio), sale e bicarbonato. "
        "2. Aggiungere acqua tiepida fino a ottenere un impasto liscio. "
        "3. Far riposare 30 minuti coperto. "
        "4. Dividere in palline e stendere sottili con il mattarello. "
        "5. Cuocere su una padella caldissima (o testo) per 2-3 minuti per lato. "
        "6. Farcire a piacere con prosciutto, squacquerone, rucola."
    ),
    "Gnocchi di ceci al pomodoro": (
        "1. Frullare i ceci lessati con un filo d'olio fino a ottenere una crema liscia. "
        "2. Aggiungere farina di ceci e un pizzico di sale, impastare. "
        "3. Formare dei piccoli gnocchi con le mani infarinate. "
        "4. Preparare il sugo: soffriggere aglio in olio, aggiungere passata di pomodoro, sale e basilico. "
        "5. Cuocere gli gnocchi in acqua bollente salata: sono pronti quando vengono a galla. "
        "6. Scolare e saltare nel sugo per un minuto. Servire con parmigiano."
    ),
    "Orecchiette broccoli e acciughe": (
        "1. Lessare le cime di broccoli in abbondante acqua salata, scolarle e tenerle da parte. "
        "2. Nella stessa acqua cuocere le orecchiette al dente. "
        "3. In una padella, scaldare olio d'oliva con aglio e acciughe sott'olio, facendole sciogliere. "
        "4. Aggiungere i broccoli e schiacchiarli leggermente con una forchetta. "
        "5. Scolare la pasta e saltarla nella padella con i broccoli. "
        "6. Aggiungere peperoncino a piacere e servire."
    ),
    "Paccheri alla crema di peperoni": (
        "1. Arrostire i peperoni al forno a 200°C per 30 minuti, pelarli e privarli dei semi. "
        "2. Frullare i peperoni con un filo d'olio, sale e un pizzico di peperoncino. "
        "3. Cuocere i paccheri in acqua bollente salata al dente. "
        "4. Scaldare la crema di peperoni in una padella ampia. "
        "5. Scolare i paccheri e saltarli nella crema di peperoni. "
        "6. Servire con basilico fresco e una grattugiata di pecorino."
    ),
    "Insalata di farro e lenticchie": (
        "1. Cuocere il farro in acqua bollente salata per circa 25 minuti, scolare e raffreddare. "
        "2. Cuocere le lenticchie in acqua per 20-25 minuti, scolare e raffreddare. "
        "3. Tagliare a cubetti pomodorini, cetrioli e cipolla rossa. "
        "4. Unire farro, lenticchie e verdure in una ciotola grande. "
        "5. Condire con olio extravergine, succo di limone, sale e pepe. "
        "6. Aggiungere prezzemolo tritato e mescolare bene. Servire a temperatura ambiente."
    ),
    "Pasta mari e monti": (
        "1. Pulire i funghi e tagliarli a fettine. Pulire i gamberi. "
        "2. Soffriggere aglio e peperoncino in olio d'oliva. "
        "3. Aggiungere i funghi e cuocere per 5 minuti a fuoco vivace. "
        "4. Aggiungere i gamberi e cuocere per 3-4 minuti. "
        "5. Sfumare con vino bianco, aggiungere pomodorini tagliati a metà. "
        "6. Cuocere la pasta al dente, scolarla e saltarla nella padella. "
        "7. Aggiungere prezzemolo tritato e servire."
    ),
    "Tortellini in brodo": (
        "1. Preparare il brodo: mettere in una pentola grande carne di manzo, gallina, sedano, carota, cipolla e sale. "
        "2. Cuocere a fuoco lento per almeno 2-3 ore, schiumando di tanto in tanto. "
        "3. Filtrare il brodo con un colino fine. "
        "4. Portare il brodo filtrato a ebollizione. "
        "5. Cuocere i tortellini nel brodo bollente per il tempo indicato sulla confezione (circa 3-5 minuti). "
        "6. Servire ben caldi con una spolverata di parmigiano reggiano."
    ),
    "insalata di pasta alle melanzane": (
        "1. Tagliare le melanzane a cubetti e grigliarle o friggerle in olio fino a doratura. "
        "2. Cuocere la pasta corta al dente, scolarla e raffreddarla sotto acqua corrente. "
        "3. Tagliare pomodorini a metà, olive a rondelle. "
        "4. Unire pasta, melanzane, pomodorini, olive e capperi. "
        "5. Condire con olio extravergine, sale, pepe e basilico fresco. "
        "6. Mescolare bene e servire a temperatura ambiente o leggermente fredda."
    ),
    "Piadina con crema di ceci e pomodorini": (
        "1. Frullare i ceci lessati con tahina, succo di limone, aglio, olio e sale fino a ottenere una crema (hummus). "
        "2. Tagliare i pomodorini a metà e condirli con olio, sale e origano. "
        "3. Scaldare la piadina su una padella calda per 1-2 minuti per lato. "
        "4. Spalmare la crema di ceci sulla piadina calda. "
        "5. Aggiungere i pomodorini, rucola e un filo d'olio. "
        "6. Piegare e servire subito."
    ),
    "risotto allo zafferano e salsiccia": (
        "1. Sciogliere lo zafferano in un mestolo di brodo caldo. "
        "2. Soffriggere cipolla tritata in burro. Aggiungere la salsiccia sbriciolata e rosolarla. "
        "3. Aggiungere il riso e tostare per 2 minuti. "
        "4. Sfumare con vino bianco. "
        "5. Aggiungere il brodo un mestolo alla volta mescolando spesso, includendo lo zafferano a metà cottura. "
        "6. Mantecare con burro e parmigiano. Servire subito."
    ),
    "farro con pomodori, feta e basilico": (
        "1. Cuocere il farro in acqua bollente salata per circa 25 minuti, scolare e raffreddare. "
        "2. Tagliare i pomodori a cubetti e la feta a dadini. "
        "3. Lavare e spezzettare il basilico fresco. "
        "4. Unire farro, pomodori, feta e basilico in una ciotola. "
        "5. Condire con olio extravergine, sale, pepe e un pizzico di origano. "
        "6. Mescolare bene e servire fresco."
    ),
    "pasta fredda con pomodorini e rucola": (
        "1. Cuocere la pasta corta al dente, scolarla e raffreddarla sotto acqua corrente. "
        "2. Tagliare i pomodorini a metà o a quarti. "
        "3. Lavare e asciugare la rucola. "
        "4. Unire pasta, pomodorini, rucola e olive (facoltativo). "
        "5. Condire con olio extravergine, sale, pepe e scaglie di parmigiano. "
        "6. Mescolare e servire fredda."
    ),
    # === SECONDI ===
    "Nuggets di pollo": (
        "1. Tagliare il petto di pollo a bocconcini regolari. "
        "2. Preparare tre ciotole: farina, uova sbattute, pangrattato con paprika e sale. "
        "3. Passare ogni bocconcino prima nella farina, poi nell'uovo, poi nel pangrattato. "
        "4. Disporre su una teglia con carta forno e spruzzare con olio d'oliva. "
        "5. Cuocere in forno a 200°C per 15-20 minuti, girando a metà cottura. "
        "6. Servire con salsa a piacere (ketchup, senape, yogurt)."
    ),
    "Tartare di carne": (
        "1. Scegliere un taglio di manzo freschissimo (filetto o fesa). Tenerlo in frigo fino all'uso. "
        "2. Con un coltello affilato, tagliare la carne a cubetti molto piccoli (non usare il tritacarne). "
        "3. Condire con olio extravergine, succo di limone, sale e pepe. "
        "4. Aggiungere capperi, cetriolini tritati e senape di Digione. "
        "5. Mescolare delicatamente e formare con un coppapasta. "
        "6. Servire subito con crostini di pane e un tuorlo d'uovo crudo sopra (facoltativo)."
    ),
    "Polpette di pollo": (
        "1. Tritare il petto di pollo al coltello o con il mixer. "
        "2. Unire uovo, pangrattato, parmigiano, prezzemolo tritato, sale e pepe. "
        "3. Impastare e formare polpette della dimensione di una noce. "
        "4. Passarle nel pangrattato. "
        "5. Cuocere in forno a 190°C per 20 minuti oppure in padella con poco olio. "
        "6. Servire con insalata o salsa di pomodoro."
    ),
    "Pollo al curry": (
        "1. Tagliare il petto di pollo a bocconcini. "
        "2. Soffriggere cipolla tritata in olio, aggiungere curry in polvere e curcuma. "
        "3. Aggiungere il pollo e rosolare per 5 minuti. "
        "4. Versare latte di cocco e un po' di brodo. "
        "5. Cuocere a fuoco medio per 15-20 minuti fino a che la salsa si addensa. "
        "6. Aggiustare di sale e servire con riso basmati."
    ),
    "pesce spada in padella": (
        "1. Condire le fette di pesce spada con olio, sale, pepe e succo di limone. "
        "2. Scaldare una padella antiaderente con un filo d'olio a fuoco vivace. "
        "3. Cuocere il pesce spada per 3-4 minuti per lato. "
        "4. Preparare un salmoriglio: olio, limone, origano, aglio tritato e prezzemolo. "
        "5. Servire il pesce spada caldo con il salmoriglio sopra. "
        "6. Accompagnare con verdure grigliate o insalata."
    ),
    "Orata al cartoccio": (
        "1. Pulire l'orata (o farla pulire dal pescivendolo). "
        "2. Farcire il ventre con fettine di limone, aglio, prezzemolo e rosmarino. "
        "3. Adagiare su un foglio di carta forno, aggiungere pomodorini, olive e capperi. "
        "4. Condire con olio, sale e pepe. Chiudere il cartoccio sigillando bene i bordi. "
        "5. Cuocere in forno a 200°C per 25-30 minuti. "
        "6. Aprire il cartoccio al momento di servire."
    ),
    "Burger di fagioli": (
        "1. Scolare e sciacquare i fagioli (cannellini o borlotti). "
        "2. Schiacciare i fagioli con una forchetta lasciando qualche pezzo intero. "
        "3. Aggiungere pangrattato, uovo, cipolla tritata, prezzemolo, cumino, sale e pepe. "
        "4. Formare dei burger e passarli nel pangrattato. "
        "5. Cuocere in padella con olio per 4-5 minuti per lato, o in forno a 190°C per 20 minuti. "
        "6. Servire in un panino con insalata, pomodoro e salsa yogurt."
    ),
    "Pollo al limone": (
        "1. Tagliare il petto di pollo a fettine sottili e infarinarle leggermente. "
        "2. Rosolare in padella con olio e burro per 3-4 minuti per lato. "
        "3. Sfumare con il succo di 2 limoni. "
        "4. Aggiungere un mestolo di brodo di pollo e cuocere per 5 minuti. "
        "5. La salsa deve diventare cremosa e lucida. "
        "6. Servire con la salsa al limone, prezzemolo fresco e fettine di limone."
    ),
    "Polpette di pesce": (
        "1. Lessare il pesce (merluzzo o nasello) e sminuzzarlo con una forchetta. "
        "2. Unire patata lessa schiacciata, uovo, pangrattato, prezzemolo, sale e pepe. "
        "3. Impastare e formare polpette. "
        "4. Passare nel pangrattato. "
        "5. Cuocere in forno a 190°C per 20 minuti oppure friggere in olio caldo. "
        "6. Servire con maionese o salsa tartara e limone."
    ),
    "Frittata di verdure": (
        "1. Tagliare le verdure di stagione (zucchine, peperoni, cipolle) a dadini. "
        "2. Saltare le verdure in padella con olio per 10 minuti. "
        "3. Sbattere le uova con sale, pepe e parmigiano grattugiato. "
        "4. Versare le uova sulle verdure nella padella. "
        "5. Cuocere a fuoco basso con coperchio per 8-10 minuti. "
        "6. Girare la frittata con un piatto e cuocere altri 3-4 minuti. Servire calda o fredda."
    ),
    # === PIATTI UNICI ===
    "Vellutata di zucca, porri e patate": (
        "1. Tagliare la zucca a cubetti, i porri a rondelle e le patate a pezzi. "
        "2. Soffriggere i porri in olio d'oliva per 5 minuti. "
        "3. Aggiungere zucca e patate, coprire con brodo vegetale. "
        "4. Cuocere a fuoco medio per 25-30 minuti fino a che le verdure sono tenere. "
        "5. Frullare con un frullatore a immersione fino a ottenere una crema liscia. "
        "6. Aggiustare di sale, pepe e servire con crostini e un filo d'olio."
    ),
    "Rotolo di melanzane": (
        "1. Tagliare le melanzane a fette lunghe e grigliarle o cuocerle in padella. "
        "2. Preparare il ripieno: ricotta, parmigiano, basilico, sale e pepe. "
        "3. Spalmare il ripieno su ogni fetta di melanzana e arrotolare. "
        "4. Disporre i rotoli in una teglia, coprire con salsa di pomodoro e mozzarella. "
        "5. Cuocere in forno a 180°C per 20-25 minuti. "
        "6. Servire caldi con basilico fresco."
    ),
    "Patate uova e avocado": (
        "1. Lessare le patate e tagliarle a cubetti. "
        "2. Cuocere le uova sode (9 minuti), pelarle e tagliarle a quarti. "
        "3. Tagliare l'avocado maturo a fette o cubetti. "
        "4. Comporre il piatto con patate, uova e avocado. "
        "5. Condire con olio extravergine, sale, pepe e succo di limone. "
        "6. Aggiungere semi di sesamo o erbe fresche a piacere."
    ),
    "Cous-cous e polpette di manzo": (
        "1. Preparare le polpette: impastare carne macinata con uovo, pangrattato, prezzemolo, sale e pepe. Formare polpette piccole. "
        "2. Rosolare le polpette in padella con olio su tutti i lati. "
        "3. Aggiungere passata di pomodoro, cuocere a fuoco basso per 15 minuti. "
        "4. Preparare il cous-cous: versare brodo bollente sul cous-cous, coprire e far riposare 5 minuti. "
        "5. Sgranare il cous-cous con una forchetta, condire con olio. "
        "6. Servire il cous-cous con le polpette e il sugo."
    ),
    "Insalata di merluzzo al vapore": (
        "1. Cuocere il merluzzo al vapore per 10-12 minuti con limone ed erbe aromatiche. "
        "2. Sfilettare il pesce e sminuzzarlo grossolanamente. "
        "3. Preparare un'insalata con lattuga, pomodorini, olive e cipolla rossa. "
        "4. Adagiare il merluzzo sull'insalata. "
        "5. Condire con vinaigrette (olio, limone, senape, sale e pepe). "
        "6. Servire a temperatura ambiente."
    ),
    "Salmone con patate, fagiolini e pomodorini": (
        "1. Tagliare le patate a cubetti e lessarle per 10 minuti. Lessare i fagiolini per 5 minuti. "
        "2. Condire il filetto di salmone con olio, sale, pepe e limone. "
        "3. Cuocere il salmone in padella per 4 minuti per lato. "
        "4. In una teglia, disporre patate, fagiolini e pomodorini tagliati a metà. Condire con olio e sale. "
        "5. Cuocere le verdure in forno a 200°C per 15 minuti. "
        "6. Servire il salmone sopra il letto di verdure."
    ),
    "Insalata di patate e salmone affumicato": (
        "1. Lessare le patate e tagliarle a cubetti una volta fredde. "
        "2. Tagliare il salmone affumicato a striscioline. "
        "3. Preparare la salsa: yogurt greco, succo di limone, aneto, sale e pepe. "
        "4. Unire patate e salmone in una ciotola. "
        "5. Condire con la salsa allo yogurt e mescolare delicatamente. "
        "6. Decorare con capperi e aneto fresco. Servire fredda."
    ),
    "Cotoletta impanata e spinaci": (
        "1. Battere le fettine di carne (pollo o vitello) con il batticarne. "
        "2. Passare nella farina, poi nell'uovo sbattuto, poi nel pangrattato. "
        "3. Cuocere in padella con olio abbondante per 3-4 minuti per lato fino a doratura. "
        "4. Scolare su carta assorbente. "
        "5. Saltare gli spinaci freschi in padella con aglio e olio per 3 minuti. "
        "6. Servire la cotoletta con gli spinaci e una spruzzata di limone."
    ),
    "Ragù di lenticchie": (
        "1. Soffriggere sedano, carota e cipolla tritati in olio d'oliva. "
        "2. Aggiungere le lenticchie (già lessate o in scatola) e mescolare. "
        "3. Versare la passata di pomodoro e un bicchiere di brodo vegetale. "
        "4. Aggiungere rosmarino, sale e pepe. "
        "5. Cuocere a fuoco basso per 25-30 minuti fino a che il ragù si addensa. "
        "6. Servire con pasta corta o polenta. Spolverare con parmigiano."
    ),
    "Polpette alla barbabietola": (
        "1. Lessare le barbabietole e schiacciarle con una forchetta. "
        "2. Unire ceci schiacciati, pangrattato, cumino, prezzemolo, sale e pepe. "
        "3. Impastare e formare polpette. "
        "4. Disporre su una teglia con carta forno e spennellare con olio. "
        "5. Cuocere in forno a 190°C per 20-25 minuti, girando a metà cottura. "
        "6. Servire con salsa allo yogurt e insalata."
    ),
    "tacchino ai ferri, riso integrale, insalata mista": (
        "1. Marinare le fettine di tacchino con olio, limone, rosmarino, sale e pepe per 15 minuti. "
        "2. Cuocere il riso integrale in acqua salata per 30-35 minuti, scolare. "
        "3. Cuocere il tacchino su una griglia o padella ben calda per 3-4 minuti per lato. "
        "4. Preparare l'insalata mista con lattuga, pomodorini, cetrioli e carote. "
        "5. Condire l'insalata con olio, aceto e sale. "
        "6. Comporre il piatto con tacchino, riso e insalata."
    ),
    "Insalata di uova, pomodorini e patate": (
        "1. Lessare le patate e tagliarle a cubetti. Cuocere le uova sode (9 minuti). "
        "2. Tagliare i pomodorini a metà e le uova a quarti. "
        "3. Unire tutti gli ingredienti in una ciotola capiente. "
        "4. Aggiungere cipolla rossa a fettine sottili e olive (facoltativo). "
        "5. Condire con olio extravergine, aceto, sale, pepe e origano. "
        "6. Mescolare delicatamente e servire a temperatura ambiente."
    ),
    "Barbabietola al forno con ravanelli, pomodorini e uova": (
        "1. Avvolgere le barbabietole in alluminio e cuocere in forno a 200°C per 45-60 minuti. "
        "2. Pelare e tagliare le barbabietole a spicchi. "
        "3. Cuocere le uova sode (9 minuti), pelarle e tagliarle a quarti. "
        "4. Tagliare i ravanelli a fettine sottili e i pomodorini a metà. "
        "5. Comporre il piatto con barbabietola, ravanelli, pomodorini e uova. "
        "6. Condire con olio, aceto balsamico, sale e pepe. Aggiungere erbe fresche."
    ),
    "Zuppa di cavolo nero e zucca": (
        "1. Tagliare la zucca a cubetti e il cavolo nero a striscioline, scartando le coste dure. "
        "2. Soffriggere cipolla e aglio in olio d'oliva. "
        "3. Aggiungere la zucca e coprire con brodo vegetale. Cuocere per 15 minuti. "
        "4. Aggiungere il cavolo nero e cuocere altri 10 minuti. "
        "5. Frullare parzialmente (lasciare qualche pezzo intero per consistenza). "
        "6. Servire con crostini di pane, olio a crudo e pepe."
    ),
    "Zuppa di fagioli e cavolo nero": (
        "1. Mettere in ammollo i fagioli secchi per 12 ore (o usare quelli in scatola). "
        "2. Soffriggere sedano, carota e cipolla tritati in olio con rosmarino. "
        "3. Aggiungere i fagioli scolati e coprire con brodo vegetale. "
        "4. Cuocere per 30 minuti (1 ora se fagioli secchi). "
        "5. Aggiungere il cavolo nero a striscioline e cuocere altri 15 minuti. "
        "6. Servire con pane tostato, olio a crudo e pepe nero."
    ),
    # === CONTORNI ===
    "Cavolofiore al forno con insalata di fagioli": (
        "1. Tagliare il cavolfiore a cimette. Condire con olio, sale, pepe e paprika. "
        "2. Cuocere in forno a 200°C per 25-30 minuti fino a doratura. "
        "3. Scolare i fagioli (cannellini), sciacquarli. "
        "4. Condire i fagioli con olio, limone, cipolla rossa tritata, prezzemolo, sale e pepe. "
        "5. Disporre il cavolfiore su un letto di insalata di fagioli. "
        "6. Servire tiepido."
    ),
    "Cavolfiore croccante": (
        "1. Tagliare il cavolfiore a cimette e lessarle per 3 minuti in acqua salata. Scolare bene. "
        "2. Preparare una pastella con farina, acqua frizzante, sale e paprika. "
        "3. Immergere le cimette nella pastella. "
        "4. Disporre su una teglia con carta forno e spruzzare con olio. "
        "5. Cuocere in forno a 220°C per 20-25 minuti fino a doratura croccante. "
        "6. Servire caldo con salsa allo yogurt o ketchup."
    ),
    "patate dolce in friggitrice": (
        "1. Pelare le patate dolci e tagliarle a bastoncini o spicchi. "
        "2. Condire con un cucchiaio di olio, sale, paprika e aglio in polvere. "
        "3. Preriscaldare la friggitrice ad aria a 200°C. "
        "4. Disporre le patate in un singolo strato nel cestello. "
        "5. Cuocere per 15-20 minuti, scuotendo il cestello a metà cottura. "
        "6. Servire calde con salsa allo yogurt o maionese."
    ),
    "ratatouille": (
        "1. Tagliare melanzane, zucchine, peperoni e pomodori a rondelle sottili. "
        "2. Preparare la base: soffriggere cipolla e aglio in olio, aggiungere passata di pomodoro, sale e pepe. "
        "3. Versare la base in una teglia rotonda. "
        "4. Alternare le rondelle di verdure in cerchi concentrici sulla base. "
        "5. Condire con olio, timo, rosmarino, sale e pepe. Coprire con alluminio. "
        "6. Cuocere in forno a 180°C per 45 minuti, poi scoprire e cuocere altri 15 minuti."
    ),
    "insalata mista": (
        "1. Lavare e asciugare lattuga, rucola o misticanza. "
        "2. Tagliare pomodorini a metà, cetrioli a rondelle e carote a julienne. "
        "3. Aggiungere cipolla rossa a fettine sottili (facoltativo). "
        "4. Disporre le verdure in un'insalatiera. "
        "5. Condire con olio extravergine, aceto (o limone), sale e pepe. "
        "6. Mescolare delicatamente e servire subito."
    ),
    "verdure grigliate": (
        "1. Tagliare zucchine, melanzane e peperoni a fette di circa 1 cm. "
        "2. Condire con olio, sale e pepe. "
        "3. Scaldare bene una griglia o piastra. "
        "4. Grigliare le verdure per 3-4 minuti per lato fino alle striature. "
        "5. Disporre su un piatto e condire con olio, aglio tritato, prezzemolo e aceto balsamico. "
        "6. Servire tiepide o a temperatura ambiente."
    ),
    "zucchine e patate al vapore": (
        "1. Pelare le patate e tagliarle a cubetti di 2 cm. Tagliare le zucchine a rondelle. "
        "2. Disporre le patate nella vaporiera, cuocere per 10 minuti. "
        "3. Aggiungere le zucchine e cuocere altri 8 minuti. "
        "4. Verificare la cottura con una forchetta: devono essere tenere ma non sfatte. "
        "5. Trasferire in una ciotola e condire con olio extravergine, sale, pepe e prezzemolo. "
        "6. Servire come contorno caldo."
    ),
    "cavolo romano al vapore": (
        "1. Lavare il cavolo romano e dividerlo in cimette. "
        "2. Disporre nella vaporiera e cuocere per 12-15 minuti fino a che è tenero. "
        "3. Nel frattempo, preparare un condimento con olio, limone, aglio tritato e sale. "
        "4. Scolare bene le cimette. "
        "5. Condire il cavolo romano con il condimento preparato. "
        "6. Servire caldo come contorno, con una spolverata di pepe."
    ),
    "Fritters di verdure": (
        "1. Grattugiare zucchine e carote, salare e far scolare per 10 minuti. Strizzare bene. "
        "2. Unire le verdure con uovo, farina, parmigiano, sale e pepe. "
        "3. Mescolare fino a ottenere un composto omogeneo. "
        "4. Formare delle frittelle con un cucchiaio. "
        "5. Cuocere in padella con olio per 3-4 minuti per lato fino a doratura. "
        "6. Scolare su carta assorbente e servire caldi."
    ),
}


def main():
    filepath = "dati/ricette/Cucina ottimizzata.xlsx"
    wb = openpyxl.load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find the column index for headers
        headers = [cell.value for cell in ws[1]]

        # Add PROCEDIMENTO header if not present
        if "PROCEDIMENTO" not in headers:
            proc_col = len(headers) + 1
            ws.cell(row=1, column=proc_col, value="PROCEDIMENTO")
        else:
            proc_col = headers.index("PROCEDIMENTO") + 1

        # Find RICETTA column
        ricetta_col = headers.index("RICETTA") + 1

        # Fill in procedures
        for row in range(2, ws.max_row + 1):
            recipe_name = ws.cell(row=row, column=ricetta_col).value
            if recipe_name:
                recipe_name = recipe_name.strip()
            if recipe_name and recipe_name in PROCEDIMENTI:
                ws.cell(row=row, column=proc_col, value=PROCEDIMENTI[recipe_name])
            elif recipe_name:
                print(f"WARNING: No procedure found for '{recipe_name}'")

    wb.save(filepath)
    print(f"File saved: {filepath}")


if __name__ == "__main__":
    main()
