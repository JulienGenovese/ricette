"""Operazioni CRUD sulle ricette nei file Excel."""

from pathlib import Path

import openpyxl
from loguru import logger

from src.excel import RecipeWriter


def remove_from_excel(recipes_dir: Path, recipe_name: str) -> bool:
    """Remove a recipe by name from all Excel files in the recipes directory.

    Returns True if the recipe was found and removed, False otherwise.
    """
    found = False
    for filepath in sorted(recipes_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(filepath)
            modified = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_to_delete = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    if row[0].value and str(row[0].value).strip() == recipe_name:
                        rows_to_delete.append(row_idx)
                # Delete in reverse to maintain row indices
                for row_idx in reversed(rows_to_delete):
                    ws.delete_rows(row_idx)
                    modified = True
                    found = True
            if modified:
                wb.save(filepath)
                logger.info("Ricetta '{}' rimossa da {}", recipe_name, filepath.name)
        except Exception as e:
            logger.warning("Errore rimuovendo '{}' da {}: {}", recipe_name, filepath.name, e)
    return found


def update_recipe_in_excel(recipes_dir: Path, original_name: str, updates: dict) -> bool:
    """Aggiorna una ricetta esistente nei file Excel.

    Returns True if found and updated, False if not found.
    """
    found = False
    for filepath in sorted(recipes_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(filepath)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=False):
                    if row[0].value and str(row[0].value).strip() == original_name:
                        # Columns: RICETTA, INGREDIENTI, STAGIONALITA, FONTE, FONTE 2
                        if updates.get("name") is not None:
                            row[0].value = updates["name"]
                        if updates.get("ingredients") is not None:
                            row[1].value = updates["ingredients"]
                        if updates.get("seasonality") is not None:
                            row[2].value = updates["seasonality"]
                        if updates.get("source1") is not None:
                            row[3].value = updates["source1"]
                        if updates.get("source2") is not None:
                            row[4].value = updates["source2"]

                        # If category changed, move to new sheet
                        if updates.get("category") and updates["category"] != sheet_name:
                            new_sheet = updates["category"]
                            if new_sheet in wb.sheetnames:
                                values = [cell.value for cell in row]
                                wb[new_sheet].append(values)
                                ws.delete_rows(row[0].row)

                        found = True
                        break
                if found:
                    break
            if found:
                wb.save(filepath)
                logger.info("Ricetta '{}' aggiornata in {}", original_name, filepath.name)
                break
        except Exception as e:
            logger.warning("Errore aggiornando '{}' in {}: {}", original_name, filepath.name, e)

    return found


def add_recipe_to_excel(
    recipes_dir: Path,
    session_file: str | None,
    timestamp: str,
    category: str,
    name: str,
    ingredients: str,
    seasonality: str,
    source1: str,
    source2: str,
) -> str:
    """Aggiunge una ricetta al file Excel della sessione corrente.

    Returns the session filename used (for tracking across calls).
    """
    expected_name = f"{timestamp}_ricette_utente.xlsx"

    # Riusa il file della sessione se esiste ancora, altrimenti crea nuovo
    if session_file:
        resolved = (recipes_dir / session_file).resolve()
        if not resolved.is_relative_to(recipes_dir.resolve()):
            raise ValueError(f"Nome file non valido: {session_file}")
    if session_file and (recipes_dir / session_file).exists():
        filepath = recipes_dir / session_file
        result_session = session_file
    else:
        filepath = recipes_dir / expected_name
        result_session = expected_name

    writer = RecipeWriter(filepath)
    writer.write_recipe(category, name, ingredients, seasonality, source1, source2)
    return result_session
