"""Moduli di export per piano settimanale e lista della spesa.

Estrae la logica di generazione Excel/PDF dal service,
eliminando la duplicazione di iterazione, scaling e styling.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from src.config import (
    DAYS_ORDER, MEALS_ORDER,
    PLAN_HEADER_COLOR, PLAN_ALT_COLOR, PLAN_EXCEL_FILLS,
    SHOPPING_HEADER_COLOR, SHOPPING_ALT_COLOR,
)


# ── Helpers condivisi ─────────────────────────────────────────


def scale_quantity(qty: float | int | None, num_people: int) -> int | None:
    """Scala una quantita' per il numero di persone, arrotondando sempre."""
    if not qty:
        return qty
    return round(qty * num_people)


def _iter_plan_rows(plan_data: dict):
    """Itera il piano in ordine giorno/pasto, yield (day, meal, meal_data)."""
    for day in DAYS_ORDER:
        day_data = plan_data.get(day)
        if not day_data:
            continue
        meals = day_data.get("meals", {})
        for meal in MEALS_ORDER:
            meal_data = meals.get(meal)
            if not meal_data:
                continue
            yield day, meal, meal_data


def _format_ingredient(ing: dict, num_people: int) -> str:
    """Formatta un ingrediente scalato come stringa."""
    scaled = scale_quantity(ing.get("quantity"), num_people)
    name = ing.get("name", "")
    unit = ing.get("unit", "") or ""
    if scaled:
        return f"{scaled}{unit} {name}" if unit else f"{scaled} {name}"
    return name


def _auto_col_widths(ws, max_width: int = 50):
    """Auto-dimensiona le colonne di un foglio Excel."""
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def _build_pdf_table(data: list[list], col_widths: list, header_color: str, alt_color: str) -> Table:
    """Costruisce una Table reportlab con stile uniforme."""
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{header_color}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{alt_color}")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def _build_pdf(title: str, table: Table, pagesize=A4) -> BytesIO:
    """Costruisce un PDF con titolo e tabella."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Spacer(1, 0.5 * cm),
        table,
    ]
    doc.build(elements)
    buf.seek(0)
    return buf


# ── Plan Export ───────────────────────────────────────────────


class PlanExporter:
    """Esporta il piano settimanale in Excel o PDF."""

    @staticmethod
    def to_excel(plan_data: dict, num_people: int) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Piano Settimanale"

        headers = ["Giorno", "Pasto", "Ricette", "Ingredienti", "Nutrienti"]
        ws.append(headers)

        bold = Font(bold=True)
        for i, cell in enumerate(ws[1]):
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            color = PLAN_EXCEL_FILLS.get(i)
            if color:
                cell.fill = PatternFill("solid", color)

        for day, meal, meal_data in _iter_plan_rows(plan_data):
            for recipe in meal_data.get("recipes", []):
                ingredients = ", ".join(
                    _format_ingredient(ing, num_people)
                    for ing in recipe.get("ingredients", [])
                )
                nutrients = ", ".join(recipe.get("nutrients", []))
                ws.append([day, meal, recipe["name"], ingredients, nutrients])

        _auto_col_widths(ws)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def to_pdf(plan_data: dict, num_people: int) -> BytesIO:
        data = [["Giorno", "Pasto", "Ricette", "Nutrienti"]]

        for day, meal, meal_data in _iter_plan_rows(plan_data):
            names = ", ".join(r["name"] for r in meal_data.get("recipes", []))
            nutrients = ", ".join(
                n for r in meal_data.get("recipes", []) for n in r.get("nutrients", [])
            )
            data.append([day, meal, names, nutrients])

        table = _build_pdf_table(
            data,
            col_widths=[2.5 * cm, 2.5 * cm, 12 * cm, 8 * cm],
            header_color=PLAN_HEADER_COLOR,
            alt_color=PLAN_ALT_COLOR,
        )
        return _build_pdf("Piano Settimanale", table, pagesize=landscape(A4))


# ── Shopping List Export ──────────────────────────────────────


class ShoppingExporter:
    """Esporta la lista della spesa in Excel o PDF."""

    @staticmethod
    def to_excel(shopping_list: list[dict], num_people: int) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista della Spesa"

        headers = ["Ingrediente", "Quantita'"]
        ws.append(headers)

        header_fill = PatternFill("solid", SHOPPING_HEADER_COLOR)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        for item in shopping_list:
            scaled = scale_quantity(item.get("quantity"), num_people)
            unit = item.get("unit", "") or ""
            qty_str = f"{scaled}{unit}" if scaled and unit else (scaled if scaled else "")
            ws.append([item["name"], qty_str])

        _auto_col_widths(ws, max_width=40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def to_pdf(shopping_list: list[dict], num_people: int) -> BytesIO:
        data = [["Ingrediente", "Quantita'"]]
        for item in shopping_list:
            scaled = scale_quantity(item.get("quantity"), num_people)
            unit = item.get("unit", "") or ""
            qty_str = f"{scaled}{unit}" if scaled and unit else (scaled if scaled else "")
            data.append([item["name"], qty_str])

        table = _build_pdf_table(
            data,
            col_widths=[12 * cm, 6 * cm],
            header_color=SHOPPING_HEADER_COLOR,
            alt_color=SHOPPING_ALT_COLOR,
        )
        return _build_pdf("Lista della Spesa", table)


# ── Dispatch map (usato da routes.py) ────────────────────────

EXPORT_HANDLERS: dict[tuple[str, str], callable] = {
    ("plan", "excel"): lambda data, np: PlanExporter.to_excel(data["plan"], np),
    ("plan", "pdf"): lambda data, np: PlanExporter.to_pdf(data["plan"], np),
    ("shopping", "excel"): lambda data, np: ShoppingExporter.to_excel(data["shopping_list"], np),
    ("shopping", "pdf"): lambda data, np: ShoppingExporter.to_pdf(data["shopping_list"], np),
}

EXPORT_MEDIA_TYPES = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}

EXPORT_FILENAMES = {
    ("plan", "excel"): "piano_settimanale.xlsx",
    ("plan", "pdf"): "piano_settimanale.pdf",
    ("shopping", "excel"): "lista_della_spesa.xlsx",
    ("shopping", "pdf"): "lista_della_spesa.pdf",
}
