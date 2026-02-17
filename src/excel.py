import os
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.model import Recipe


class RecipeRepository:
    CATEGORIES = ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]

    def __init__(self, dirpath: str | Path):
        self._dirpath = Path(dirpath)

    def list_files(self) -> list[str]:
        """Ritorna i nomi dei file .xlsx disponibili nella directory."""
        return sorted(f.name for f in self._dirpath.glob("*.xlsx"))

    def load_recipes(self) -> tuple[dict[str, list[Recipe]], set[str], dict[str, str]]:
        """Carica ricette da tutti i file .xlsx nella directory.
        Ritorna: ({categoria: [lista_ricette]}, set_nomi_ricette_utente, {nome_ricetta: nome_file})
        """
        recipes_by_category: dict[str, list[Recipe]] = {cat: [] for cat in self.CATEGORIES}
        user_recipe_names: set[str] = set()
        recipe_sources: dict[str, str] = {}
        required_cols = {"RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"}

        xlsx_files = sorted(self._dirpath.glob("*.xlsx"))
        if not xlsx_files:
            logger.warning("Nessun file .xlsx trovato in {}", self._dirpath)
            return recipes_by_category, user_recipe_names, recipe_sources

        for filepath in xlsx_files:
            is_user_file = "_ricette_utente.xlsx" in filepath.name
            logger.info("Caricamento ricette da: {} (utente: {})", filepath.name, is_user_file)
            for category in self.CATEGORIES:
                try:
                    df = pd.read_excel(filepath, sheet_name=category)

                    if not required_cols.issubset(df.columns):
                        raise ValueError(f"Colonne mancanti nel sheet {category}")

                    recipes = [
                        Recipe(
                            name=row["RICETTA"].strip() if isinstance(row["RICETTA"], str) else row["RICETTA"],
                            ingredients=row["INGREDIENTI"],
                            seasonality=row["STAGIONALITA"],
                            source1=row["FONTE"],
                            source2=row["FONTE 2"],
                        )
                        for _, row in df.iterrows()
                    ]
                    recipes_by_category[category].extend(recipes)
                    for r in recipes:
                        recipe_sources[r.name] = filepath.name
                    if is_user_file:
                        user_recipe_names.update(r.name for r in recipes)
                except Exception:
                    logger.debug("Sheet '{}' non trovato o non valido in {}", category, filepath.name)

        for cat, recipes in recipes_by_category.items():
            logger.info("{}: {} ricette caricate", cat, len(recipes))

        return recipes_by_category, user_recipe_names, recipe_sources


class BaseExcelWriter:
    def __init__(self, output_path: str, sheet_name: str):
        self.output_path = output_path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

        self.header_font = Font(bold=True)
        self.font = Font(bold=False)

        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_left = Alignment(horizontal="left", vertical="center")

    def save(self):
        dirpath = os.path.dirname(self.output_path)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
        self.wb.save(self.output_path)


class PlanExcelWriter(BaseExcelWriter):
    def __init__(self, output_path: str):
        super().__init__(output_path, "Piano Settimanale")

        self.fills = {
            "Giorno": PatternFill("solid", "BDD7EE"),
            "Pasto": PatternFill("solid", "C6E0B4"),
            "Ricette": PatternFill("solid", "F8CBAD"),
            "Nutrienti": PatternFill("solid", "FFF2CC"),
        }

    def write(self, plan):
        self.ws.append(["Giorno", "Pasto", "Ricette", "Nutrienti"])

        for day, meals in plan.items():
            for meal, data in meals.items():
                if isinstance(data, dict):
                    recipes = data.get("recipes", [])
                    nutrients = data.get("nutrients", "")
                else:
                    recipes = data
                    nutrients = ""

                self.ws.append([day, meal, ", ".join(r.name for r in recipes), nutrients])

        self._format()
        self.save()

    def _format(self):
        for col in self.ws.iter_cols():
            header = col[0].value
            fill = self.fills.get(header)

            for i, cell in enumerate(col):
                cell.font = self.header_font if i == 0 else self.font
                if header in ("Ricette", "Nutrienti"):
                    cell.alignment = self.align_left
                else:
                    cell.alignment = self.align_center
                if i == 0 and fill:
                    cell.fill = fill


class ShoppingListWriter(BaseExcelWriter):
    def __init__(self, output_path: str):
        super().__init__(output_path, "Lista della Spesa")

    def write(self, shopping_list, aggregations=None, units=None):
        self.ws.append(["Ingrediente", "Quantita'", "Varianti Aggregate"])

        for cell in self.ws[1]:
            cell.font = self.header_font
            cell.alignment = self.align_center

        units = units or {}
        for ingredient, qty in sorted(shopping_list.items()):
            variants = ", ".join(
                v for v in aggregations.get(ingredient, []) if v != ingredient
            ) if aggregations else ""

            unit = units.get(ingredient, "")
            if qty:
                qty_str = f"{qty}{unit}" if unit else str(qty)
            else:
                qty_str = ""

            self.ws.append([
                ingredient,
                qty_str,
                variants
            ])

        self.save()


class RecipeWriter:
    """Scrive/appende ricette in un file Excel con il formato standard.

    Il file ha 4 sheet (PRIMI, SECONDI, PIATTI UNICI, CONTORNI)
    con colonne: RICETTA, INGREDIENTI, STAGIONALITA, FONTE, FONTE 2.
    """

    CATEGORIES = RecipeRepository.CATEGORIES
    COLUMNS = ["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"]

    def __init__(self, filepath: str | Path):
        self._filepath = Path(filepath)

    def write_recipe(self, category: str, name: str, ingredients: str,
                     seasonality: str, source1: str, source2: str):
        if category not in self.CATEGORIES:
            raise ValueError(f"Categoria non valida: {category}")

        if self._filepath.exists():
            wb = load_workbook(self._filepath)
        else:
            wb = Workbook()
            wb.remove(wb.active)
            for cat in self.CATEGORIES:
                ws = wb.create_sheet(cat)
                ws.append(self.COLUMNS)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws = wb[category]
        ws.append([name, ingredients, seasonality, source1, source2])

        os.makedirs(self._filepath.parent, exist_ok=True)
        wb.save(self._filepath)
        logger.info("Ricetta '{}' salvata in {} (sheet {})", name, self._filepath.name, category)
