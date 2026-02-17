"""Test per la feature di priorità/probabilità delle ricette."""
import random
from collections import Counter
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from src.model import Recipe, CategorizedRecipe, DishType
from src.optimizer import MealBuilder, classify_recipes
from src.rules import MealRuleEngine, NutritionRule, PiattoUnicoRule


# ───────── Helpers ─────────

def _recipe(name, source1="Proteica", source2="Fibra"):
    return Recipe(
        name=name,
        ingredients="100g pasta",
        seasonality="Tutto l'anno",
        source1=source1,
        source2=source2,
    )


def _categorized(name, dish_type=DishType.PRIMO, source1="Proteica", source2="Fibra"):
    r = _recipe(name, source1, source2)
    return CategorizedRecipe(recipe=r, dish_type=dish_type)


def _profiles_for(categorized_recipes):
    """Genera profili nutrizionali per una lista di CategorizedRecipe."""
    from src.model import NutritionClassifier
    classifier = NutritionClassifier()
    return {cr.recipe: classifier.classify(cr.recipe) for cr in categorized_recipes}


# ═══════════════════════════════════════════
#  Test RecipeRepository: tracciamento file utente
# ═══════════════════════════════════════════

class TestRecipeRepositoryUserFiles:
    """Verifica che load_recipes() identifichi correttamente i file utente."""

    def test_load_recipes_returns_tuple(self, tmp_path):
        """load_recipes deve ritornare (dict, set, dict)."""
        from src.excel import RecipeRepository
        repo = RecipeRepository(tmp_path)
        result = repo.load_recipes()

        assert isinstance(result, tuple)
        assert len(result) == 3
        recipes, user_names, recipe_sources = result
        assert isinstance(recipes, dict)
        assert isinstance(user_names, set)
        assert isinstance(recipe_sources, dict)

    def test_user_file_names_detected(self, tmp_path):
        """Ricette da file *_ricette_utente.xlsx devono finire nel set."""
        import openpyxl

        # Creo un file utente finto
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
            if cat == "PRIMI":
                ws.append(["Pasta utente", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "2026-01-01_10-00_ricette_utente.xlsx")

        from src.excel import RecipeRepository
        repo = RecipeRepository(tmp_path)
        _, user_names, _ = repo.load_recipes()

        assert "Pasta utente" in user_names

    def test_non_user_file_not_in_set(self, tmp_path):
        """Ricette da file normali NON devono finire nel set utente."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
            if cat == "PRIMI":
                ws.append(["Pasta base", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "Cucina ottimizzata.xlsx")

        from src.excel import RecipeRepository
        repo = RecipeRepository(tmp_path)
        _, user_names, _ = repo.load_recipes()

        assert "Pasta base" not in user_names


# ═══════════════════════════════════════════
#  Test MealBuilder: weighted sampling
# ═══════════════════════════════════════════

class TestMealBuilderWeights:
    """Verifica che MealBuilder utilizzi i pesi nel sampling."""

    @pytest.fixture
    def recipes_and_profiles(self):
        """Crea un set di ricette con profili per i test."""
        recipes = [
            _categorized("Ricetta_A", source1="Proteica", source2="Fibra"),
            _categorized("Ricetta_B", source1="Proteica", source2="Fibra"),
            _categorized("Ricetta_C", source1="Proteica", source2="Fibra"),
            _categorized("Ricetta_D", source1="Proteica", source2="Fibra"),
            _categorized("Ricetta_E", source1="Proteica", source2="Fibra"),
        ]
        profiles = _profiles_for(recipes)
        return recipes, profiles

    def test_builder_accepts_weights(self, recipes_and_profiles):
        """MealBuilder deve accettare il parametro weights."""
        recipes, profiles = recipes_and_profiles
        weights = {"Ricetta_A": 5.0}
        rules = MealRuleEngine([NutritionRule()])
        builder = MealBuilder(rules, profiles, weights)

        assert builder.weights == weights

    def test_builder_default_weights_empty(self, recipes_and_profiles):
        """Senza pesi, weights deve essere un dict vuoto."""
        _, profiles = recipes_and_profiles
        rules = MealRuleEngine([NutritionRule()])
        builder = MealBuilder(rules, profiles)

        assert builder.weights == {}

    def test_high_weight_recipe_appears_more_often(self, recipes_and_profiles):
        """Una ricetta con peso molto alto deve apparire più spesso delle altre."""
        recipes, profiles = recipes_and_profiles

        # Diamo un peso enorme a Ricetta_A
        weights = {
            "Ricetta_A": 100.0,
            "Ricetta_B": 1.0,
            "Ricetta_C": 1.0,
            "Ricetta_D": 1.0,
            "Ricetta_E": 1.0,
        }

        rules = MealRuleEngine([NutritionRule()])
        builder = MealBuilder(rules, profiles, weights)

        counts = Counter()
        n_trials = 200
        random.seed(42)

        for _ in range(n_trials):
            meal = builder.build(recipes)
            for cr in meal:
                counts[cr.recipe.name] += 1

        # Ricetta_A con peso 100x deve apparire significativamente più delle altre
        assert counts["Ricetta_A"] > counts.get("Ricetta_B", 0)
        assert counts["Ricetta_A"] > counts.get("Ricetta_C", 0)

    def test_equal_weights_still_works(self, recipes_and_profiles):
        """Con pesi uguali, il builder deve comunque funzionare."""
        recipes, profiles = recipes_and_profiles
        weights = {r.recipe.name: 1.0 for r in recipes}

        rules = MealRuleEngine([NutritionRule()])
        builder = MealBuilder(rules, profiles, weights)

        random.seed(42)
        meal = builder.build(recipes)

        assert len(meal) >= 1
        assert all(isinstance(cr, CategorizedRecipe) for cr in meal)

    def test_no_weights_still_works(self, recipes_and_profiles):
        """Senza pesi (backward compatibility), il builder deve funzionare."""
        recipes, profiles = recipes_and_profiles

        rules = MealRuleEngine([NutritionRule()])
        builder = MealBuilder(rules, profiles)  # nessun weights

        random.seed(42)
        meal = builder.build(recipes)

        assert len(meal) >= 1


# ═══════════════════════════════════════════
#  Test RecipeService: gestione pesi
# ═══════════════════════════════════════════

class TestRecipeServiceWeights:
    """Verifica che RecipeService gestisca correttamente i pesi."""

    @pytest.fixture
    def service(self):
        from api.service import RecipeService
        return RecipeService()

    def test_initial_weights_empty(self, service):
        """I pesi iniziali devono essere vuoti prima del load."""
        assert service._weights == {}

    def test_load_initializes_all_weights_to_one(self, service, tmp_path):
        """Dopo load(), tutte le ricette devono avere peso 1.0 (senza file utente)."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        ws_primi = wb["PRIMI"]
        ws_primi.append(["Pasta", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        ws_primi.append(["Riso", "100g riso", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        with patch("api.service.RECIPES_DIR", tmp_path):
            service.load()

        assert service._weights["Pasta"] == 1.0
        assert service._weights["Riso"] == 1.0

    def test_user_file_recipes_get_boost(self, service, tmp_path):
        """Ricette da file utente devono ricevere il boost USER_FILE_BOOST."""
        import openpyxl

        # File base
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        wb["PRIMI"].append(["Pasta base", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        # File utente
        wb2 = openpyxl.Workbook()
        wb2.remove(wb2.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb2.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        wb2["PRIMI"].append(["Pasta utente", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb2.save(tmp_path / "2026-01-01_10-00_ricette_utente.xlsx")

        with patch("api.service.RECIPES_DIR", tmp_path):
            service.load()

        assert service._weights["Pasta base"] == 1.0
        assert service._weights["Pasta utente"] == 1.0 + service.USER_FILE_BOOST

    def test_featured_recipes_boost_weights(self, service, tmp_path):
        """get_featured_recipes() deve aumentare il peso delle ricette selezionate."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        wb["PRIMI"].append(["R1", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb["PRIMI"].append(["R2", "100g riso", "Tutto l'anno", "Proteica", "Fibra"])
        wb["SECONDI"].append(["R3", "200g pollo", "Tutto l'anno", "Proteica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        with patch("api.service.RECIPES_DIR", tmp_path):
            service.load()

        # Tutti a 1.0 prima
        for name in ["R1", "R2", "R3"]:
            assert service._weights[name] == 1.0

        featured = service.get_featured_recipes(n=3)

        # Le ricette in vetrina non devono modificare i pesi (side-effect rimosso)
        featured_names = {r["name"] for r in featured}
        for name in featured_names:
            assert service._weights[name] == 1.0

    def test_featured_recipes_returns_correct_format(self, service, tmp_path):
        """get_featured_recipes() deve ritornare il formato corretto."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        wb["PRIMI"].append(["Pasta test", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        with patch("api.service.RECIPES_DIR", tmp_path):
            service.load()

        featured = service.get_featured_recipes(n=1)

        assert len(featured) == 1
        item = featured[0]
        assert "name" in item
        assert "dish_type" in item
        assert "nutrients" in item
        assert isinstance(item["nutrients"], list)

    def test_featured_does_not_modify_weights(self, service, tmp_path):
        """get_featured_recipes() non deve modificare i pesi (nessun side-effect)."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        wb["PRIMI"].append(["Unica", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        with patch("api.service.RECIPES_DIR", tmp_path):
            service.load()

        service.get_featured_recipes(n=1)
        service.get_featured_recipes(n=1)

        assert service._weights["Unica"] == 1.0

    def test_featured_respects_n_limit(self, service, tmp_path):
        """get_featured_recipes(n) non deve ritornare più di n ricette."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        for i in range(10):
            wb["PRIMI"].append([f"R{i}", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        with patch("api.service.RECIPES_DIR", tmp_path):
            service.load()

        featured = service.get_featured_recipes(n=3)
        assert len(featured) == 3


# ═══════════════════════════════════════════
#  Test API endpoint: /api/featured-recipes
# ═══════════════════════════════════════════

class TestFeaturedRecipesEndpoint:
    """Verifica il funzionamento dell'endpoint featured-recipes."""

    @pytest.fixture
    def client(self, tmp_path):
        """Crea un test client FastAPI con ricette di test."""
        import openpyxl
        from contextlib import asynccontextmanager
        from fastapi import FastAPI
        from fastapi.testclient import TestClient
        from api.routes import router
        from api.service import RecipeService

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cat in ["PRIMI", "SECONDI", "PIATTI UNICI", "CONTORNI"]:
            ws = wb.create_sheet(cat)
            ws.append(["RICETTA", "INGREDIENTI", "STAGIONALITA", "FONTE", "FONTE 2"])
        wb["PRIMI"].append(["Test Pasta", "100g pasta", "Tutto l'anno", "Glucidica", "Fibra"])
        wb["SECONDI"].append(["Test Pollo", "200g pollo", "Tutto l'anno", "Proteica", "Fibra"])
        wb.save(tmp_path / "base.xlsx")

        @asynccontextmanager
        async def lifespan(app: FastAPI):
            service = RecipeService()
            with patch("api.service.RECIPES_DIR", tmp_path):
                service.load()
            app.state.recipe_service = service
            yield

        app = FastAPI(lifespan=lifespan)
        app.include_router(router, prefix="/api")

        with TestClient(app) as c:
            yield c

    def test_featured_endpoint_returns_200(self, client):
        """L'endpoint deve ritornare 200."""
        resp = client.get("/api/featured-recipes")
        assert resp.status_code == 200

    def test_featured_endpoint_returns_recipes(self, client):
        """L'endpoint deve ritornare una lista di ricette."""
        resp = client.get("/api/featured-recipes")
        data = resp.json()

        assert "recipes" in data
        assert isinstance(data["recipes"], list)
        assert len(data["recipes"]) > 0

    def test_featured_recipe_has_required_fields(self, client):
        """Ogni ricetta deve avere name, dish_type, nutrients."""
        resp = client.get("/api/featured-recipes")
        data = resp.json()

        for recipe in data["recipes"]:
            assert "name" in recipe
            assert "dish_type" in recipe
            assert "nutrients" in recipe
